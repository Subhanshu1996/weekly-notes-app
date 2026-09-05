import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
import io
import uuid
import os
import html
import re
import smtplib
from email.message import EmailMessage
import hashlib
import random
from collections import Counter
import gspread
from google.oauth2.service_account import Credentials

# --- Email Logo Fallback URL ---
EMAIL_LOGO_URL = "https://www.factspan.com/wp-content/uploads/2021/10/Factspan-Logo.png"

# Optional PDF dependency
try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False

# Optional Excel styling dependency
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# =========================================================
# APP CONFIGURATION & TIMEZONE
# =========================================================
st.set_page_config(page_title="Weekly Project Report", page_icon="📈", layout="wide")
IST = ZoneInfo("Asia/Kolkata")

# =========================================================
# GOOGLE SHEETS BACKEND HELPERS
# =========================================================
@st.cache_resource
def get_gspread_client():
    scopes = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
    return gspread.authorize(creds)

def update_user_password_in_gsheets(email, new_hashed_pwd):
    try:
        client = get_gspread_client()
        spreadsheet = client.open("Weekly Notes Database")
        u_sheet = spreadsheet.worksheet("Users")
        records = u_sheet.get_all_records()
        headers = u_sheet.row_values(1)
        if "Password_Hash" not in headers: return False
        col_idx = headers.index("Password_Hash") + 1
        for idx, row in enumerate(records):
            if str(row.get("Email", "")).strip().lower() == email.strip().lower():
                row_idx = idx + 2 
                u_sheet.update_cell(row_idx, col_idx, new_hashed_pwd)
                return True
        return False
    except Exception as e: return False

def send_otp_email(recipient_email, otp, context="Password Reset"):
    cfg = st.secrets if hasattr(st, "secrets") else {}
    host = cfg.get("SMTP_HOST", os.getenv("SMTP_HOST", ""))
    port = int(cfg.get("SMTP_PORT", os.getenv("SMTP_PORT", "587")))
    username = cfg.get("SMTP_USERNAME", os.getenv("SMTP_USERNAME", ""))
    password = cfg.get("SMTP_PASSWORD", os.getenv("SMTP_PASSWORD", ""))
    sender = cfg.get("SMTP_FROM", os.getenv("SMTP_FROM", username))
    use_tls = str(cfg.get("SMTP_USE_TLS", os.getenv("SMTP_USE_TLS", "true"))).lower() == "true"
    
    if not host or not sender: return False, "Email sending is not configured in secrets."
    
    msg = EmailMessage()
    msg["Subject"] = f"{context} - Weekly Project Report"
    msg["From"] = sender
    msg["To"] = recipient_email
    msg.set_content(f"Your One-Time Password (OTP) to verify your account is: {otp}\n\nIf you did not request this, please ignore this email.")
    
    try:
        with smtplib.SMTP(host, port, timeout=20) as server:
            if use_tls: server.starttls()
            if username and password: server.login(username, password)
            server.send_message(msg)
        return True, "OTP sent successfully."
    except Exception as exc: return False, f"Unable to send email: {exc}"

# =========================================================
# SECURITY & ROLE-BASED LOGIN / REGISTRATION SCREEN (RESTORED)
# =========================================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.current_email = ""
    st.session_state.current_name = ""
    st.session_state.current_role = ""
    st.session_state.current_scope = "" 
    st.session_state.current_acc_scope = ""
    st.session_state.reset_otp = None
    st.session_state.reset_email = None
    st.session_state.reg_otp = None
    st.session_state.reg_details = None

if not st.session_state.authenticated:
    st.markdown('<div style="max-width: 480px; margin: 60px auto; padding: 30px; background: white; border-radius: 12px; border: 1px solid #dfe5ec; box-shadow: 0 4px 12px rgba(0,0,0,0.1);">', unsafe_allow_html=True)
    st.markdown("<h2 style='color: #1e293b; text-align:center; margin-bottom: 5px;'>🔒 Project Portal</h2>", unsafe_allow_html=True)
    st.markdown("<p style='color: #64748b; margin-bottom: 20px; text-align:center;'>Log in or create an account to continue.</p>", unsafe_allow_html=True)
    
    tab1, tab2, tab3 = st.tabs(["🔐 Log In", "📝 Create Account", "🔑 Forgot Password"])
    
    with tab1:
        log_email = st.text_input("Company Email", key="log_email", placeholder="name@company.com").strip().lower()
        log_pwd = st.text_input("Password", type="password", key="log_pwd", placeholder="Your personal password")
        
        if st.button("Log In", type="primary", use_container_width=True):
            if not log_email or not log_pwd: st.error("Please enter your email and password.")
            else:
                with st.spinner("Authenticating..."):
                    client = get_gspread_client()
                    spreadsheet = client.open("Weekly Notes Database")
                    try: u_sheet = spreadsheet.worksheet("Users"); records = u_sheet.get_all_records()
                    except Exception: records = []
                    
                    user = next((r for r in records if str(r.get("Email", "")).strip().lower() == log_email), None)
                    hashed_attempt = hashlib.sha256(log_pwd.encode()).hexdigest()
                    
                    if user and user.get("Password_Hash") == hashed_attempt:
                        st.session_state.authenticated = True
                        st.session_state.current_email = str(user.get("Email", ""))
                        st.session_state.current_name = str(user.get("Name", ""))
                        st.session_state.current_role = str(user.get("Role", ""))
                        st.session_state.current_scope = str(user.get("Scope", ""))
                        st.session_state.current_acc_scope = str(user.get("Account_Scope", ""))
                        st.rerun()
                    else: st.error("Invalid email or password.")
                        
    with tab2:
        if not st.session_state.reg_otp:
            reg_name = st.text_input("Full Name", key="reg_name", placeholder="e.g. John Smith").strip()
            reg_email = st.text_input("Company Email", key="reg_email", placeholder="name@company.com").strip().lower()
            reg_role = st.selectbox("Select Your Role", ["Team Member", "Team Lead", "Account Manager", "Admin"], key="reg_role")
            
            reg_acc_scope, reg_scope = "", ""
            if reg_role == "Team Lead":
                reg_acc_scope = st.text_input("Account Name", placeholder="e.g. Acme Corp", key="reg_acc_scope").strip()
                reg_scope = st.text_input("Your Team(s)", placeholder="e.g. Alpha Team, Beta Team", key="reg_scope").strip()
            elif reg_role == "Account Manager":
                reg_scope = st.text_input("Your Account(s)", placeholder="e.g. Acme Corp, Globex", key="reg_scope2").strip()
                
            c1, c2 = st.columns(2)
            with c1: reg_pwd = st.text_input("Create Password", type="password", key="reg_pwd")
            with c2: reg_pwd2 = st.text_input("Confirm Password", type="password", key="reg_pwd2")
            st.divider()
            
            reg_invite = ""
            if reg_role != "Team Member":
                reg_invite = st.text_input(f"{reg_role} Invite Code", type="password", placeholder="Enter secret code...", key="reg_invite")
            else: st.caption("✨ Team Members can register directly without an invite code.")
            
            if st.button("Send Verification Code", type="primary", use_container_width=True):
                if not reg_name: st.error("Please enter your name.")
                elif not reg_email or not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", reg_email): st.error("Please enter a valid company email address.")
                elif reg_role == "Team Lead" and (not reg_acc_scope or not reg_scope): st.error("Please enter both your Account and Team scope.")
                elif reg_role == "Account Manager" and not reg_scope: st.error("Please enter your Account(s) scope.")
                elif len(reg_pwd) < 6: st.error("Password must be at least 6 characters.")
                elif reg_pwd != reg_pwd2: st.error("Passwords do not match.")
                else:
                    invite_valid = False
                    if reg_role == "Team Member": invite_valid = True
                    else:
                        password_keys = {"Team Lead": "LEAD_PASSWORD", "Account Manager": "ACCOUNT_PASSWORD", "Admin": "ADMIN_PASSWORD"}
                        req_invite = st.secrets.get(password_keys[reg_role], os.getenv(password_keys[reg_role], ""))
                        if not req_invite: st.error(f"Configuration error: {password_keys[reg_role]} is not set.")
                        elif reg_invite != req_invite: st.error(f"Invalid Invite Code for the {reg_role} role.")
                        else: invite_valid = True
                    
                    if invite_valid:
                        with st.spinner("Checking details..."):
                            client = get_gspread_client()
                            spreadsheet = client.open("Weekly Notes Database")
                            try: u_sheet = spreadsheet.worksheet("Users"); records = u_sheet.get_all_records()
                            except Exception: records = []
                            
                            if any(str(r.get("Email","")).strip().lower() == reg_email for r in records):
                                st.error("An account with this email already exists. Please log in.")
                            else:
                                otp = str(random.randint(100000, 999999))
                                ok, msg = send_otp_email(reg_email, otp, context="Account Verification")
                                if ok:
                                    st.session_state.reg_otp = otp
                                    st.session_state.reg_details = {"name": reg_name, "email": reg_email, "role": reg_role, "scope": reg_scope, "acc_scope": reg_acc_scope, "pwd": reg_pwd}
                                    st.rerun()
                                else: st.error(msg)
        else:
            st.success(f"Verification code sent to {st.session_state.reg_details['email']}")
            reg_entered_otp = st.text_input("Enter 6-digit OTP", key="reg_entered_otp", placeholder="000000")
            st.divider()
            rc1, rc2 = st.columns(2)
            with rc1:
                if st.button("Verify & Create Account", type="primary", use_container_width=True):
                    if reg_entered_otp.strip() != st.session_state.reg_otp: st.error("Invalid verification code.")
                    else:
                        with st.spinner("Creating your account securely..."):
                            client = get_gspread_client()
                            spreadsheet = client.open("Weekly Notes Database")
                            try: u_sheet = spreadsheet.worksheet("Users")
                            except Exception:
                                u_sheet = spreadsheet.add_worksheet(title="Users", rows=1000, cols=6)
                                u_sheet.append_row(["Email", "Name", "Role", "Scope", "Account_Scope", "Password_Hash"])
                            hashed_pw = hashlib.sha256(st.session_state.reg_details['pwd'].encode()).hexdigest()
                            u_sheet.append_row([st.session_state.reg_details['email'], st.session_state.reg_details['name'], st.session_state.reg_details['role'], st.session_state.reg_details['scope'], st.session_state.reg_details['acc_scope'], hashed_pw])
                            st.session_state.authenticated = True
                            st.session_state.current_email = st.session_state.reg_details['email']
                            st.session_state.current_name = st.session_state.reg_details['name']
                            st.session_state.current_role = st.session_state.reg_details['role']
                            st.session_state.current_scope = st.session_state.reg_details['scope']
                            st.session_state.current_acc_scope = st.session_state.reg_details['acc_scope']
                            st.session_state.reg_otp, st.session_state.reg_details = None, None
                            st.rerun()
            with rc2:
                if st.button("Cancel", use_container_width=True):
                    st.session_state.reg_otp, st.session_state.reg_details = None, None
                    st.rerun()

    with tab3:
        if not st.session_state.reset_otp:
            st.markdown("<p style='font-size:14px; color:#475569;'>Enter your registered email address to receive a secure 6-digit reset code.</p>", unsafe_allow_html=True)
            reset_email_input = st.text_input("Registered Company Email", key="reset_email_in", placeholder="name@company.com").strip().lower()
            if st.button("Send OTP", use_container_width=True):
                if not reset_email_input: st.error("Please enter your email.")
                else:
                    with st.spinner("Verifying account..."):
                        client = get_gspread_client()
                        spreadsheet = client.open("Weekly Notes Database")
                        try: u_sheet = spreadsheet.worksheet("Users"); records = u_sheet.get_all_records()
                        except Exception: records = []
                        user = next((r for r in records if str(r.get("Email", "")).strip().lower() == reset_email_input), None)
                        if not user: st.error("No account found with this email.")
                        else:
                            otp = str(random.randint(100000, 999999))
                            ok, msg = send_otp_email(reset_email_input, otp, context="Password Reset")
                            if ok:
                                st.session_state.reset_otp = otp
                                st.session_state.reset_email = reset_email_input
                                st.success(f"An OTP has been securely sent to {reset_email_input}")
                                st.rerun()
                            else: st.error(msg)
        else:
            st.success(f"OTP sent to {st.session_state.reset_email}")
            entered_otp = st.text_input("Enter 6-digit OTP", key="entered_otp", placeholder="000000")
            st.divider()
            new_pwd1 = st.text_input("New Password", type="password", key="new_pwd1", placeholder="At least 6 characters")
            new_pwd2 = st.text_input("Confirm New Password", type="password", key="new_pwd2")
            rc1, rc2 = st.columns(2)
            with rc1:
                if st.button("Reset Password", type="primary", use_container_width=True):
                    if entered_otp.strip() != st.session_state.reset_otp: st.error("Invalid OTP.")
                    elif len(new_pwd1) < 6: st.error("Password must be at least 6 characters.")
                    elif new_pwd1 != new_pwd2: st.error("Passwords do not match.")
                    else:
                        with st.spinner("Updating password securely..."):
                            hashed_pw = hashlib.sha256(new_pwd1.encode()).hexdigest()
                            success = update_user_password_in_gsheets(st.session_state.reset_email, hashed_pw)
                            if success:
                                st.success("Password updated successfully! Please switch to the Log In tab.")
                                st.session_state.reset_otp, st.session_state.reset_email = None, None
                            else: st.error("Failed to update password. Please try again.")
            with rc2:
                if st.button("Cancel Reset", use_container_width=True):
                    st.session_state.reset_otp, st.session_state.reset_email = None, None
                    st.rerun()

    st.markdown('</div>', unsafe_allow_html=True)
    st.stop()


def load_data_from_gsheets():
    try:
        client = get_gspread_client()
        sheet = client.open("Weekly Notes Database").sheet1
        return sheet.get_all_records()
    except Exception as e:
        st.error(f"Failed to connect to Google Sheets. Error: {e}")
        return []

def save_new_notes_to_gsheets(new_entries):
    client = get_gspread_client()
    sheet = client.open("Weekly Notes Database").sheet1
    headers = sheet.row_values(1)
    if "Resource Name" not in headers:
        headers.append("Resource Name")
        sheet.update_cell(1, len(headers), "Resource Name")
    rows_to_append = [[str(entry.get(col, "")) for col in headers] for entry in new_entries]
    sheet.append_rows(rows_to_append)

def update_existing_note_in_gsheets(updated_entry):
    client = get_gspread_client()
    sheet = client.open("Weekly Notes Database").sheet1
    records = sheet.get_all_records()
    headers = sheet.row_values(1)
    if "Resource Name" not in headers:
        headers.append("Resource Name")
        sheet.update_cell(1, len(headers), "Resource Name")
    for idx, row in enumerate(records):
        if str(row.get("id")) == str(updated_entry.get("id")):
            row_idx = idx + 2 
            update_values = [[str(updated_entry.get(col, "")) for col in headers]]
            sheet.update(f"A{row_idx}:{get_column_letter(len(headers))}{row_idx}", update_values)
            break

# =========================================================
# PREMIUM UI SYSTEM
# =========================================================
st.markdown("""
<style>
:root {
    --app-bg: #f4f7fb; --surface: #ffffff; --primary: #1e293b; --accent: #ea580c; --border: #dfe5ec;
    --text-main: #0f172a; --muted: #64748b; --radius-sm: 8px; --radius-md: 12px;
}
.stApp { background: var(--app-bg); }
.block-container { max-width: 1440px !important; margin: 0 auto !important; padding-top: 1.5rem !important; padding-bottom: 2.5rem !important; }
.stMarkdown, .stText, .stCaption, .stSelectbox label, .stMultiSelect label, .stTextInput label, .stTextArea label, .stNumberInput label, .stDateInput label { font-family: Inter, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif !important; }
.stSelectbox label p, .stMultiSelect label p, .stTextInput label p, .stTextArea label p, .stNumberInput label p, .stDateInput label p, .stSelectbox label, .stMultiSelect label, .stTextInput label, .stTextArea label, .stNumberInput label, .stDateInput label { font-size: 15px !important; font-weight: bold !important; color: #1e293b !important; letter-spacing: .25px !important; margin-bottom: 4px !important; }
[data-baseweb="select"] > div { background: linear-gradient(to bottom, #ffffff, #f8fafc) !important; border: 1px solid #cbd5e1 !important; box-shadow: 0 2px 4px rgba(0,0,0,0.03), 0 1px 2px rgba(0,0,0,0.04) !important; border-radius: 8px !important; min-height: 38px !important; transition: all 0.15s ease !important; }
[data-baseweb="select"] > div:hover { border-color: #94a3b8 !important; box-shadow: 0 4px 6px rgba(0,0,0,0.06), 0 2px 4px rgba(0,0,0,0.05) !important; }
.stSelectbox [data-baseweb="select"] span, .stMultiSelect [data-baseweb="select"] span { font-size: 15px !important; }
.stMultiSelect [data-baseweb="tag"] { font-size: 14px !important; font-weight: bold !important; }
.app-hero { background: linear-gradient(135deg, #1e293b 0%, #334155 100%); border-radius: 18px; padding: 28px 32px; min-height: 110px; box-sizing: border-box; color: #ffffff !important; box-shadow: 0 12px 30px rgba(30, 41, 59, .18); margin-bottom: 16px; position: relative; overflow: hidden; isolation: isolate; border-top: 4px solid #ea580c; }
.app-hero:after { content: ""; position: absolute; width: 260px; height: 260px; right: -80px; top: -120px; border-radius: 50%; background: rgba(255,255,255,.08); }
.app-hero > * { position: relative; z-index: 2; }
.app-hero-eyebrow { font-size: 13px; font-weight: bold; letter-spacing: 1.7px; text-transform: uppercase; color: #cbd5e1; margin-bottom: 6px; }
.app-hero-title { font-size: 34px; line-height: 1.15; font-weight: bold; letter-spacing: -.6px; margin: 0; color: #fff !important; }
.app-hero-subtitle { margin-top: 8px; font-size: 15px; color: #f8fafc !important; }
.stButton > button, .stDownloadButton > button { background: linear-gradient(to bottom, #ffffff, #f1f5f9) !important; border: 1px solid #cbd5e1 !important; box-shadow: 0 3px 0 #cbd5e1, 0 4px 6px rgba(0,0,0,0.05) !important; border-radius: 10px !important; color: #334155 !important; font-weight: bold !important; transition: all 0.1s ease-out !important; transform: translateY(0) !important; }
.stButton > button:active, .stDownloadButton > button:active { transform: translateY(3px) !important; box-shadow: 0 0px 0 transparent, 0 0px 0px rgba(0,0,0,0) !important; }
.stButton > button[kind="primary"] { background: linear-gradient(to bottom, var(--accent), #c2410c) !important; border-color: #9a3412 !important; box-shadow: 0 3px 0 #9a3412, 0 5px 12px rgba(234, 88, 12, 0.3) !important; color: white !important; }
div.element-container:has(.marker-nav1-inactive) + div.element-container button { background: linear-gradient(to bottom, #f8fafc, #eff6ff) !important; border-color: #bfdbfe !important; box-shadow: 0 3px 0 #bfdbfe, 0 4px 6px rgba(0,0,0,0.05) !important; color: #1e40af !important; }
div.element-container:has(.marker-nav1-active) + div.element-container button { background: linear-gradient(to bottom, #1e293b, #0f172a) !important; border-color: #020617 !important; box-shadow: 0 3px 0 #020617, 0 4px 8px rgba(15,23,42,0.3) !important; color: #ffffff !important; }
div.element-container:has(.marker-nav2-inactive) + div.element-container button { background: linear-gradient(to bottom, #f8fafc, #f0fdf4) !important; border-color: #bbf7d0 !important; box-shadow: 0 3px 0 #bbf7d0, 0 4px 6px rgba(0,0,0,0.05) !important; color: #166534 !important; }
div.element-container:has(.marker-nav2-active) + div.element-container button { background: linear-gradient(to bottom, #10b981, #047857) !important; border-color: #064e3b !important; box-shadow: 0 3px 0 #064e3b, 0 4px 8px rgba(16,185,129,0.3) !important; color: #ffffff !important; }
div.element-container:has(.marker-nav3-inactive) + div.element-container button { background: linear-gradient(to bottom, #f8fafc, #fffbeb) !important; border-color: #fde68a !important; box-shadow: 0 3px 0 #fde68a, 0 4px 6px rgba(0,0,0,0.05) !important; color: #b45309 !important; }
div.element-container:has(.marker-nav3-active) + div.element-container button { background: linear-gradient(to bottom, #ea580c, #c2410c) !important; border-color: #9a3412 !important; box-shadow: 0 3px 0 #9a3412, 0 4px 8px rgba(234,88,12,0.3) !important; color: #ffffff !important; }
div.element-container:has(.marker-nav4-inactive) + div.element-container button { background: linear-gradient(to bottom, #f8fafc, #fff1f2) !important; border-color: #fecdd3 !important; box-shadow: 0 3px 0 #fecdd3, 0 4px 6px rgba(0,0,0,0.05) !important; color: #9f1239 !important; }
div.element-container:has(.marker-nav4-active) + div.element-container button { background: linear-gradient(to bottom, #ef4444, #b91c1c) !important; border-color: #7f1d1d !important; box-shadow: 0 3px 0 #7f1d1d, 0 4px 8px rgba(239,68,68,0.3) !important; color: #ffffff !important; }
div.element-container:has(.marker-send) + div.element-container button { background: linear-gradient(to bottom, #f5f3ff, #ede9fe) !important; border-color: #c4b5fd !important; box-shadow: 0 3px 0 #c4b5fd, 0 4px 6px rgba(0,0,0,0.05) !important; color: #5b21b6 !important; }
div.element-container:has(.marker-pdf) + div.element-container button { background: linear-gradient(to bottom, #fff1f2, #ffe4e6) !important; border-color: #fecdd3 !important; box-shadow: 0 3px 0 #fecdd3, 0 4px 6px rgba(0,0,0,0.05) !important; color: #9f1239 !important; }
div.element-container:has(.marker-excel) + div.element-container button { background: linear-gradient(to bottom, #ecfdf5, #d1fae5) !important; border-color: #a7f3d0 !important; box-shadow: 0 3px 0 #a7f3d0, 0 4px 6px rgba(0,0,0,0.05) !important; color: #065f46 !important; }
div.element-container:has([class^="marker-"]) + div.element-container button:active { transform: translateY(3px) !important; box-shadow: 0 0px 0 transparent, 0 0px 0px rgba(0,0,0,0) !important; }
.page-head { background: var(--surface); border: 1px solid var(--border); border-radius: 14px; padding: 16px 20px; box-shadow: var(--shadow-sm); margin-bottom: 15px; min-height: 80px; display: flex; flex-direction: column; justify-content: center; }
.page-title { font-size: 28px; line-height: 1.2; font-weight: bold; color: var(--primary) !important; margin: 0; }
.page-subtitle { color: var(--muted) !important; font-size: 15px; margin-top: 5px; }
.section-header { font-size: 22px; font-weight: bold; color: var(--primary) !important; margin: 18px 0 10px 0; display: flex; align-items: center; gap: 8px; }
.section-header:before { content: ""; width: 5px; height: 22px; background: var(--accent); border-radius: 3px; }
[data-testid="stVerticalBlockBorderWrapper"] { border-radius: 12px !important; border-color: var(--border) !important; box-shadow: var(--shadow-sm) !important; background: var(--surface) !important; }
.card-title { font-size: 17px; font-weight: bold; color: var(--primary) !important; margin-bottom: 3px; }
.kpi-card { background: var(--surface); border: 1px solid var(--border); border-radius: 13px; padding: 15px 18px; box-shadow: var(--shadow-sm); min-height: 85px; border-top: 4px solid var(--accent); transition: transform .18s ease, box-shadow .18s ease; }
.kpi-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.kpi-label { font-size: 12px; color: var(--muted) !important; font-weight: bold; letter-spacing: .8px; }
.kpi-value { margin-top: 5px; font-size: 32px; line-height: 1; font-weight: bold; color: var(--primary) !important; }
.custom-table { width: 100%; border-collapse: separate; border-spacing: 0; background: #fff; border: 1px solid var(--border); border-radius: 12px; overflow: hidden; font-size: 15px; box-shadow: var(--shadow-sm); }
.custom-table th { background: #eef3f8; color: #344054; padding: 13px 16px; text-align: left; font-size: 13.5px; font-weight: bold; text-transform: uppercase; letter-spacing: .45px; border-bottom: 1px solid var(--border); }
.custom-table td { padding: 14px 16px; border-bottom: 1px solid #edf1f5; color: #344054; vertical-align: middle; }
.custom-table tr:hover td { background: #f8fbff; }
.proj-name { font-size: 15px; font-weight: bold; color: var(--primary); display:block; }
.work-type { font-size: 13px; color: var(--muted); display:block; margin-top:2px; }
.status-pill { display: inline-flex; align-items: center; padding: 5px 9px; border-radius: 999px; font-size: 12px; font-weight: bold; border: 1px solid transparent; }
.status-on-track { background:#ecfdf5; color:#16805b; border-color:#a7f3d0; }
.status-in-progress { background:#fffbeb; color:#ea580c; border-color:#fde68a; }
.status-blocked { background:#fef2f2; color:#c2413d; border-color:#fecaca; }
.status-completed { background:#eff6ff; color:#2563eb; border-color:#bfdbfe; }
.status-not-started { background:#f2f4f7; color:#667085; border-color:#d0d5dd; }
.summary-box { background: #f8fafc; border-radius: 8px; color: #334155; font-size: 1.15rem; line-height: 1.6; margin-bottom: 2.5rem; padding: 1.25rem; border: 1px solid #e2e8f0; }
.risk-box { background-color: #fff7ed; border-left: 5px solid #ea580c; padding: 1.25rem; margin-bottom: 1rem; border-radius: 6px; }
.risk-title { color: #0f172a; font-size: 1.05rem; font-weight: 700; margin-bottom: 0.25rem; }
.risk-desc { color: #334155; font-size: 0.95rem; }
.acc-header { background: linear-gradient(90deg,#1e293b,#334155); color:#fff !important; padding:9px 14px; border-radius:10px; font-size:17px; font-weight:bold; margin:13px 0 7px 0; }
.team-header { color:var(--primary) !important; font-size:15px; font-weight:bold; padding:5px 2px; border-bottom:2px solid #e6ebf1; margin:7px 0 5px; }
.edit-banner { background:#eff6ff; color:#2563eb !important; border:1px solid #bfdbfe; padding:8px 11px; border-radius:9px; font-size:13px; font-weight:bold; margin-bottom:10px; }
.block-container .stVerticalBlock { gap: .8rem !important; }
.action-critical { background:#fef2f2; border:1px solid #fecaca; border-left:5px solid #dc2626; padding:12px 14px; border-radius:10px; margin:8px 0; }
.action-attention { background:#fffbeb; border:1px solid #fde68a; border-left:5px solid #d97706; padding:12px 14px; border-radius:10px; margin:8px 0; }
hr { margin: 12px 0 !important; border-color:#e6ebf1 !important; }
.filter-shell { background:#fff; border:1px solid #dfe5ec; border-radius:14px; padding:10px 14px 4px; margin:8px 0 14px; box-shadow:0 2px 8px rgba(15,23,42,.04); }
.filter-caption { color:#64748b; font-size:12px; margin-top:-2px; }
.filter-summary { display:flex; flex-wrap:wrap; gap:6px; align-items:center; margin:2px 0 8px; }
.filter-chip { display:inline-block; background:#fff7ed; color:#ea580c; border:1px solid #fdba74; border-radius:999px; padding:4px 9px; font-size:12px; font-weight:700; }
.health-strip { background:#fff; border:1px solid #dfe5ec; border-radius:14px; padding:12px 15px; margin:8px 0 15px; }
.health-title { font-size:12px; text-transform:uppercase; letter-spacing:.8px; color:#64748b; font-weight:800; }
.health-score { font-size:25px; font-weight:800; color:#1e293b; }
.health-dot { display:inline-block; width:9px; height:9px; border-radius:50%; margin-right:6px; }
.mini-insight { background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:10px 12px; font-size:13px; color:#334155; }
.login-meta { text-align:center; color:#94a3b8; font-size:11px; margin-top:8px; }
.section-note { color:#64748b; font-size:12px; margin-top:-5px; margin-bottom:8px; }
.compact-row { border-bottom:1px solid #edf1f5; padding:9px 4px; }
</style>
""", unsafe_allow_html=True)

# =========================================================
# SESSION STATE & LIVE DATA SYNC
# =========================================================
if 'data_synced' not in st.session_state:
    st.session_state.notes_db = load_data_from_gsheets()
    st.session_state.data_synced = True

if 'project_count' not in st.session_state: st.session_state.project_count = 1
if 'show_success' not in st.session_state: st.session_state.show_success = False
if 'success_message' not in st.session_state: st.session_state.success_message = ""
if 'edit_id' not in st.session_state: st.session_state.edit_id = None
if 'nav_selection' not in st.session_state: st.session_state.nav_selection = "📋 Weekly Summary"
if 'show_send_dialog' not in st.session_state: st.session_state.show_send_dialog = None
if 'last_refresh_at' not in st.session_state: st.session_state.last_refresh_at = datetime.now(IST)
if 'quick_filter' not in st.session_state: st.session_state.quick_filter = "All"
if 'project_search' not in st.session_state: st.session_state.project_search = ""
if 'executive_compact' not in st.session_state: st.session_state.executive_compact = False
if 'copy_source_loaded' not in st.session_state: st.session_state.copy_source_loaded = False

st.session_state.accounts_db = {}
if st.session_state.notes_db:
    live_df = pd.DataFrame(st.session_state.notes_db)
    if "Account" in live_df.columns and "Team" in live_df.columns:
        for acc in live_df["Account"].dropna().unique():
            st.session_state.accounts_db[acc] = live_df[live_df["Account"] == acc]["Team"].unique().tolist()

# =========================================================
# NEW CENTRALIZED REPORT ENGINE
# =========================================================

def _report_period_label(dataframe):
    if dataframe is None or dataframe.empty: return "Current Selection"
    vals = []
    for col in ["Year", "Month", "Week"]:
        if col in dataframe.columns:
            unique = [str(v) for v in dataframe[col].dropna().unique() if not is_blank(v)]
            if len(unique) == 1: vals.append(unique[0])
    return " · ".join(vals) if vals else "Current Selection"

def _report_scope_text(dataframe):
    if dataframe is None or dataframe.empty: return "No records in the current selection"
    accounts = dataframe["Account"].nunique() if "Account" in dataframe.columns else 0
    teams = dataframe["Team"].nunique() if "Team" in dataframe.columns else 0
    return f"{len(dataframe)} update(s) · {accounts} account(s) · {teams} team(s)"

def _risk_dataframe(dataframe):
    if dataframe is None or dataframe.empty: return pd.DataFrame()
    status = dataframe.get("Status", pd.Series(index=dataframe.index, dtype=object)).fillna("").astype(str).str.strip()
    blocker = dataframe.get("Blocker", pd.Series(index=dataframe.index, dtype=object)).fillna("").astype(str).str.strip()
    return dataframe.loc[status.isin(["At Risk", "Blocked", "On Hold"]) | blocker.ne("")].copy()

def _delivery_outlook(dataframe, as_of=None):
    as_of = as_of or datetime.now(IST).date()
    result = {"Overdue": 0, "Next 7 Days": 0, "8-14 Days": 0, "15+ Days": 0, "Completed": 0, "No Date": 0}
    if dataframe is None or dataframe.empty: return result
    for _, row in dataframe.iterrows():
        status = normalize_text(row.get("Status"))
        if status == "completed": result["Completed"] += 1; continue
        due = date_from_row(row)
        if not due: result["No Date"] += 1; continue
        days = (due - as_of).days
        if days < 0: result["Overdue"] += 1
        elif days <= 7: result["Next 7 Days"] += 1
        elif days <= 14: result["8-14 Days"] += 1
        else: result["15+ Days"] += 1
    return result

def _avg_completion(dataframe):
    if dataframe is None or dataframe.empty or "Completion %" not in dataframe.columns: return 0
    return int(round(dataframe["Completion %"].apply(pct_value).mean()))

def _avg_utilization(dataframe):
    if dataframe is None or dataframe.empty or "Utilization %" not in dataframe.columns: return 0
    work = dataframe.assign(_util=dataframe["Utilization %"].apply(pct_value))
    if "Resource" in work.columns:
        values = work.groupby("Resource")["_util"].mean()
        return int(round(values.mean())) if not values.empty else 0
    return int(round(work["_util"].mean()))

def _portfolio_health(dataframe):
    if dataframe is None or dataframe.empty: return 0
    total = max(len(dataframe), 1)
    risks = len(_risk_dataframe(dataframe))
    overdue = _delivery_outlook(dataframe)["Overdue"]
    score = 70 + (_avg_completion(dataframe) * 0.25) - (risks / total * 18) - (overdue / total * 22)
    return int(max(0, min(100, round(score))))

def _health_label(score):
    return "Healthy" if score >= 80 else "Attention" if score >= 60 else "Critical"

def _leadership_takeaway(dataframe, movement_df=None):
    if dataframe is None or dataframe.empty: return "No records are available for the selected reporting period."
    total = len(dataframe); resources = dataframe["Resource"].nunique() if "Resource" in dataframe.columns else 0
    risks = len(_risk_dataframe(dataframe)); out = _delivery_outlook(dataframe); health = _portfolio_health(dataframe)
    text = f"Portfolio health is {health}/100 ({_health_label(health).lower()}) across {total} update(s) and {resources} resource(s), with average initiative completion at {_avg_completion(dataframe)}%. "
    text += f"{risks} item(s) require risk or blocker attention. " if risks else "No active risk/blocker items are currently flagged. "
    if out["Overdue"] or out["Next 7 Days"]: text += f"Delivery outlook includes {out['Overdue']} overdue and {out['Next 7 Days']} item(s) due within 7 days. "
    if movement_df is not None and not movement_df.empty:
        improved, declined = int((movement_df["Movement"] == "Improved").sum()), int((movement_df["Movement"] == "Declined").sum())
        if improved or declined: text += f"WoW movement shows {improved} improved and {declined} declined item(s)."
    return text

def _build_report_model(dataframe, report_title):
    data = dataframe.copy() if dataframe is not None else pd.DataFrame()
    prior_df = pd.DataFrame()
    if not data.empty and {"Year", "Month", "Week"}.issubset(data.columns):
        ys, ms, ws = [data[c].dropna().unique().tolist() for c in ["Year", "Month", "Week"]]
        if len(ys) == len(ms) == len(ws) == 1: prior_df = get_latest_prior_week(data, ys[0], ms[0], ws[0])
    movement_df = compare_current_to_prior(data, prior_df)
    actions = build_action_items(data); risks_df = _risk_dataframe(data)
    updates = int(data["This Week Delivered"].fillna("").astype(str).str.strip().ne("").sum()) if "This Week Delivered" in data.columns else 0
    kpis = {
        "projects": len(data), "resources": int(data["Resource"].nunique()) if not data.empty and "Resource" in data.columns else 0,
        "updates": updates, "completion": _avg_completion(data), "utilization": _avg_utilization(data),
        "health": _portfolio_health(data), "critical": sum(x["Priority"] == "Critical" for x in actions),
        "attention": sum(x["Priority"] == "Attention" for x in actions), "risks": len(risks_df)
    }
    mc = {k: int((movement_df["Movement"] == k).sum()) for k in ["New", "Improved", "Declined", "Stable"]} if not movement_df.empty else {k: 0 for k in ["New", "Improved", "Declined", "Stable"]}
    highlights, priorities = [], []
    for _, row in data.iterrows():
        project = str(row.get("Project / Dashboard", "")).strip()
        delivered = str(row.get("This Week Delivered", "")).strip()
        nxt = str(row.get("Next Week Priority", "")).strip()
        if project and delivered and delivered.lower() not in {"none", "none reported", "n/a"}: highlights.append((project, delivered))
        if project and nxt and nxt.lower() not in {"none", "none reported", "n/a"}: priorities.append((project, nxt))
    
    accounts = list(data["Account"].dropna().unique()) if "Account" in data else []
    teams = list(data["Team"].dropna().unique()) if "Team" in data else []
    acc_str = accounts[0] if len(accounts) == 1 else "Portfolio"
    team_str = teams[0] if len(teams) == 1 else "Summary"
    account_team_str = f"{acc_str} - {team_str}"
    
    return {
        "title": report_title, "account_team_str": account_team_str, "period": _report_period_label(data), "scope": _report_scope_text(data),
        "generated": datetime.now(IST), "data": data, "prior": prior_df, "movement": movement_df,
        "actions": actions, "risks_df": risks_df, "outlook": _delivery_outlook(data), "kpis": kpis,
        "movement_counts": mc, "health_label": _health_label(kpis["health"]),
        "takeaway": _leadership_takeaway(data, movement_df), "highlights": highlights[:10], "priorities": priorities[:10]
    }

# =========================================================
# EMAIL HTML REPORT GENERATOR
# =========================================================
def _email_status_class(status):
    s = normalize_text(status)
    if s == "completed": return "success"
    if s == "blocked": return "critical"
    if s in {"at risk", "on hold"}: return "warning"
    if s == "on track": return "info"
    return "neutral"

def generate_html_report(report):
    k, m, o = report["kpis"], report["movement_counts"], report["outlook"]
    data = report["data"]
    esc = lambda v: html_escape(v)
    
    logo_src = "cid:factspan_logo" if os.path.exists("logo.png") else EMAIL_LOGO_URL
    logo_html = f'<img src="{logo_src}" alt="FACTSPAN" width="200" style="display: block; border: 0; max-width: 100%; height: auto;" />'

    rows = []
    for _, r in data.head(20).iterrows():
        status = str(r.get("Status", ""))
        comp = int(pct_value(r.get("Completion %")))
        status_color = "#16a34a" if normalize_text(status) == "completed" else ("#ea580c" if normalize_text(status) in ["at risk", "in progress", "on hold"] else ("#dc2626" if normalize_text(status) == "blocked" else "#64748b"))
        
        raw_name = str(r.get("Resource Name", "")).strip()
        if raw_name and raw_name.lower() not in ["none", "na", "n/a", "nan", ""]: res = raw_name
        else: res = str(r.get("Resource", "")).strip().split('@')[0].replace('.', ' ').title()

        rows.append(f'''
        <tr>
            <td style="padding: 14px 12px; border-bottom: 1px solid #e2e8f0; font-size: 13px; color: #1e293b; word-break: break-word;">
                <strong>{esc(r.get("Project / Dashboard", ""))}</strong><br><span style="color: #64748b; font-size: 11px;">Owner: {esc(r.get("Business Owner", ""))}</span>
            </td>
            <td style="padding: 14px 12px; border-bottom: 1px solid #e2e8f0; font-size: 13px; color: #475569; word-break: break-word;">{esc(res)}</td>
            <td style="padding: 14px 12px; border-bottom: 1px solid #e2e8f0; vertical-align: middle;">
                <div style="font-size: 11px; font-weight: bold; color: #1e293b; margin-bottom: 4px;">{comp}%</div>
                <div style="background-color: #e2e8f0; width: 100%; height: 6px; border-radius: 3px; overflow: hidden;">
                    <div style="background-color: {status_color}; width: {comp}%; height: 100%;"></div>
                </div>
            </td>
            <td style="padding: 14px 12px; border-bottom: 1px solid #e2e8f0; font-size: 13px; color: #475569;">{esc(r.get("Updated Expected Delivery date", ""))}</td>
            <td style="padding: 14px 12px; border-bottom: 1px solid #e2e8f0; font-size: 12px; font-weight: bold; color: {status_color};">{esc(status)}</td>
        </tr>
        ''')
        
    table_html = "".join(rows) if rows else '<tr><td colspan="5" style="padding: 14px 12px; text-align: center; color: #64748b;">No records in current selection.</td></tr>'

    actions = []
    for item in report["actions"][:6]:
        critical = item["Priority"] == "Critical"
        action = "Escalate dependency and confirm recovery plan." if critical else "Confirm owner commitment and next milestone."
        actions.append(f'<li style="margin-bottom: 8px;"><span style="color: #ea580c;">■</span> <strong>{esc(item["Project / Dashboard"])}:</strong> {esc(item["Reason"])}. <em>Action: {esc(action)}</em></li>')
    action_html = "".join(actions) if actions else '<li><span style="color: #ea580c;">■</span> Routine progress tracking across all initiatives. No critical blocks reported.</li>'

    return f'''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
            @media only screen and (max-width: 600px) {{
                .email-container {{ width: 100% !important; padding: 0 !important; }}
                .mobile-stack {{ display: block !important; width: 100% !important; padding: 10px 0 !important; text-align: center !important; }}
                .mobile-hide {{ display: none !important; }}
                .kpi-box {{ display: block !important; width: 100% !important; margin-bottom: 10px !important; box-sizing: border-box; }}
                .logo-img {{ margin: 0 auto !important; }}
                .table-scroll {{ overflow-x: auto !important; display: block !important; width: 100% !important; }}
                .responsive-table {{ min-width: 600px !important; }}
                .footer-text {{ text-align: center !important; padding: 5px 0 !important; display: block !important; width: 100% !important; }}
            }}
        </style>
    </head>
    <body style="margin: 0; padding: 20px; background-color: #f4f7fb;">
        <div class="email-container" style="font-family: Arial, sans-serif; max-width: 850px; margin: 0 auto; background: #ffffff; border: 1px solid #dfe5ec;">
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="padding: 25px;">
                <tr>
                    <td class="mobile-stack" width="40%" valign="middle" align="left">
                        {logo_html}
                    </td>
                    <td class="mobile-stack" width="60%" align="right" valign="middle">
                        <div style="color: #ea580c; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.5px;">WEEKLY PROJECT REPORT</div>
                        <div style="font-size: 22px; color: #1e293b; margin: 4px 0 2px 0; font-weight: bold;">{esc(report["account_team_str"])}</div>
                        <div style="color: #64748b; font-size: 12px;">Weekly Progress Snapshot &bull; {report["generated"].strftime('%d %b %Y')}</div>
                    </td>
                </tr>
            </table>
            <div style="border-top: 4px solid #ea580c;"></div>
            <div style="padding: 25px;">
                <p style="color: #334155; font-size: 14px; margin-top: 0;">Hi Team,</p>
                <p style="color: #475569; font-size: 14px; line-height: 1.6; margin-bottom: 25px;">
                    Please find below the latest weekly progress snapshot for <strong>{esc(report["account_team_str"])}</strong>.
                </p>
                <h3 style="color: #0f172a; font-size: 16px; margin-bottom: 12px; border-bottom: 2px solid #ea580c; padding-bottom: 5px; display: inline-block;">Portfolio Snapshot</h3>
                
                <table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom: 25px;">
                    <tr>
                        <td class="kpi-box" width="23%" align="center" style="border: 1px solid #e2e8f0; padding: 20px 0; background: #f8fafc;">
                            <div style="font-size: 28px; font-weight: bold; color: #1e293b; font-family: Arial, sans-serif;">{k["projects"]}</div>
                            <div style="font-size: 10px; color: #64748b; font-family: Arial, sans-serif; text-transform: uppercase; margin-top: 4px;">PROJECTS</div>
                        </td>
                        <td class="mobile-hide" width="2%"></td>
                        <td class="kpi-box" width="23%" align="center" style="border: 1px solid #e2e8f0; padding: 20px 0; background: #f8fafc;">
                            <div style="font-size: 28px; font-weight: bold; color: #16a34a; font-family: Arial, sans-serif;">{k["updates"]}</div>
                            <div style="font-size: 10px; color: #64748b; font-family: Arial, sans-serif; text-transform: uppercase; margin-top: 4px;">UPDATES</div>
                        </td>
                        <td class="mobile-hide" width="2%"></td>
                        <td class="kpi-box" width="23%" align="center" style="border: 1px solid #e2e8f0; padding: 20px 0; background: #fffbeb;">
                            <div style="font-size: 28px; font-weight: bold; color: #ea580c; font-family: Arial, sans-serif;">{k["completion"]}%</div>
                            <div style="font-size: 10px; color: #64748b; font-family: Arial, sans-serif; text-transform: uppercase; margin-top: 4px;">COMPLETION</div>
                        </td>
                        <td class="mobile-hide" width="2%"></td>
                        <td class="kpi-box" width="23%" align="center" style="border: 1px solid #e2e8f0; padding: 20px 0; background: #fef2f2;">
                            <div style="font-size: 28px; font-weight: bold; color: #dc2626; font-family: Arial, sans-serif;">{k["risks"]}</div>
                            <div style="font-size: 10px; color: #64748b; font-family: Arial, sans-serif; text-transform: uppercase; margin-top: 4px;">RISKS</div>
                        </td>
                    </tr>
                </table>
                
                <div class="table-scroll">
                    <table class="responsive-table" width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse: collapse; border: 1px solid #e2e8f0; margin-bottom: 30px;">
                        <thead>
                            <tr style="background-color: #1e293b;">
                                <th align="left" width="30%" style="padding: 14px 12px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase;">PROJECT / OWNER</th>
                                <th align="left" width="20%" style="padding: 14px 12px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase;">RESOURCE</th>
                                <th align="left" width="25%" style="padding: 14px 12px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase;">DELIVERY PROGRESS</th>
                                <th align="left" width="15%" style="padding: 14px 12px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase;">TARGET DATE</th>
                                <th align="left" width="10%" style="padding: 14px 12px; color: #ffffff; font-size: 11px; font-weight: bold; text-transform: uppercase;">STATUS</th>
                            </tr>
                        </thead>
                        <tbody>
                            {table_html}
                        </tbody>
                    </table>
                </div>
                
                <div style="background-color: #fff7ed; border-left: 4px solid #ea580c; padding: 20px; margin-bottom: 30px;">
                    <h4 style="color: #0f172a; margin: 0 0 12px 0; font-size: 16px;">Key Updates & Actions</h4>
                    <ul style="margin: 0; padding-left: 20px; color: #475569; font-size: 13px; line-height: 1.6;">
                        {action_html}
                    </ul>
                </div>
            </div>
            <table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color: #1e293b; padding: 15px 25px;">
                <tr>
                    <td class="footer-text" align="left" style="color: #f1f5f9; font-size: 11px;">
                        Factspan &bull; {esc(report["account_team_str"])}
                    </td>
                    <td class="footer-text" align="right" style="color: #ea580c; font-size: 11px; font-weight: bold;">
                        Staying Relevant and Ahead through Fluid Intelligence
                    </td>
                </tr>
            </table>
        </div>
    </body>
    </html>
    '''

# =========================================================
# PDF REPORT GENERATOR
# =========================================================
def _pdf_safe(value):
    text = "" if value is None else str(value)
    for a,b in {"–":"-","—":"-","’":"'","“":'"',"”":'"',"•":"-","✓":"OK","⚠":"!","→":"->","₹":"Rs ","…":"..."}.items(): text=text.replace(a,b)
    return text.encode("latin-1","replace").decode("latin-1")

def _pdf_section(pdf, title):
    y = pdf.get_y() + 3
    pdf.set_fill_color(37, 99, 235); pdf.rect(15, y+1, 2.2, 7, "F")
    pdf.set_xy(19, y); pdf.set_text_color(22, 50, 79); pdf.set_font("Arial", "B", 13); pdf.cell(170, 8, _pdf_safe(title), ln=True); pdf.ln(1)

def _pdf_lines(pdf, text, width, size=8.2, max_lines=4):
    text=_pdf_safe(text); words=text.split(); pdf.set_font("Arial","",size)
    if not words: return [""]
    lines=[]; cur=""
    for word in words:
        candidate=word if not cur else cur+" "+word
        if pdf.get_string_width(candidate)<=width: cur=candidate
        else:
            lines.append(cur or word); cur=word
            if len(lines)>=max_lines: break
    if cur and len(lines)<max_lines: lines.append(cur)
    if len(lines)==max_lines and len(" ".join(lines))<len(text):
        last=lines[-1]
        while last and pdf.get_string_width(last+"...")>width: last=last.rsplit(" ",1)[0] if " " in last else last[:-1]
        lines[-1]=(last.rstrip(" .")+"...") if last else "..."
    return lines

def _pdf_table(pdf, dataframe, headers, widths, getter):
    if dataframe is None or dataframe.empty: return
    pdf.set_x(15)
    pdf.set_fill_color(231, 238, 246); pdf.set_text_color(37, 54, 74); pdf.set_font("Arial", "B", 7.8)
    for h,w in zip(headers,widths): pdf.cell(w,8,_pdf_safe(h),border=1,fill=True)
    pdf.ln()
    for ridx,(_,row) in enumerate(dataframe.iterrows()):
        vals=getter(row); line_sets=[_pdf_lines(pdf,v,w-3) for v,w in zip(vals,widths)]
        h=max(8,max(len(x) for x in line_sets)*4.5+2)
        if pdf.get_y()+h>270:
            pdf.add_page(); pdf.set_x(15); pdf.set_y(20); pdf.set_fill_color(231, 238, 246); pdf.set_text_color(37, 54, 74); pdf.set_font("Arial", "B", 7.8)
            for hh,ww in zip(headers,widths): pdf.cell(ww,8,_pdf_safe(hh),border=1,fill=True)
            pdf.ln()
        x=15;y=pdf.get_y();pdf.set_fill_color(249,251,253) if ridx%2 else pdf.set_fill_color(255,255,255)
        for lines,w,val in zip(line_sets,widths,vals):
            pdf.rect(x,y,w,h,"DF"); pdf.set_xy(x+1.5,y+1.5)
            
            if "Risk" in val or "Blocked" in val: pdf.set_text_color(160, 55, 55)
            elif "Completed" in val: pdf.set_text_color(22, 128, 91)
            else: pdf.set_text_color(52, 64, 84)
            
            pdf.set_font("Arial","",8.2)
            for line in lines:
                pdf.cell(w-3,4.5,_pdf_safe(line),ln=True);pdf.set_x(x+1.5)
            x+=w
        pdf.set_y(y+h)

def create_pdf_report(dataframe, header_title):
    if not FPDF_AVAILABLE: return None
    report=_build_report_model(dataframe,header_title);k=report["kpis"];o=report["outlook"]
    
    class ReportPDF(FPDF):
        def header(self):
            if os.path.exists("logo.png"):
                try: self.image("logo.png", 15, 10, 40)
                except Exception:
                    self.set_text_color(30, 41, 59); self.set_font("Arial", "B", 16)
                    self.set_xy(15, 15); self.cell(40, 10, "FACTSPAN", ln=False)
            else:
                self.set_text_color(30, 41, 59); self.set_font("Arial", "B", 16)
                self.set_xy(15, 15); self.cell(40, 10, "FACTSPAN", ln=False)

            self.set_text_color(234, 88, 12)
            self.set_font("Arial", "B", 8)
            self.set_xy(15, 10)
            self.cell(180, 4, "WEEKLY PROJECT REPORT", align="R", ln=True)
            
            self.set_text_color(30, 41, 59)
            self.set_font("Arial", "B", 18)
            self.set_x(15)
            self.cell(180, 8, safe_pdf_text(report["account_team_str"]), align="R", ln=True)
            
            self.set_text_color(100, 116, 139)
            self.set_font("Arial", "", 8)
            self.set_x(15)
            self.cell(180, 4, safe_pdf_text(header_title), align="R", ln=True)
            
            self.set_x(15)
            self.cell(180, 4, safe_pdf_text(f"Generated: {report['generated'].strftime('%d %b %Y, %I:%M %p')}"), align="R", ln=True)
            
            self.set_y(32)
            self.set_draw_color(234, 88, 12)
            self.set_line_width(1)
            self.line(15, 32, 195, 32)
            self.set_line_width(0.2)
            self.ln(5) 

        def footer(self):
            self.set_y(-15)
            self.set_draw_color(234, 88, 12)
            self.set_line_width(0.5)
            self.line(15, 282, 195, 282)
            self.set_text_color(30, 41, 59)
            self.set_font("Arial", "B", 8)
            self.set_xy(15, 285)
            self.cell(90, 5, safe_pdf_text(f"Factspan \x95 {report['account_team_str']}"), align="L")
            self.set_text_color(100, 116, 139)
            self.set_font("Arial", "", 8)
            self.set_xy(100, 285)
            self.cell(95, 5, f"Page {self.page_no()}", align="R")

    pdf=ReportPDF("P","mm","A4"); pdf.set_margins(15,15,15); pdf.set_auto_page_break(True,18); pdf.add_page()
    
    pdf.set_y(40)
    cards=[("PROJECTS",k["projects"]),("RESOURCES",k["resources"]),("UPDATES",k["updates"]),("COMPLETION",f"{k['completion']}%"),("RISKS",k["risks"])]
    box_w = 33.6; gap = 3; y = 40
    for i,(lab,val) in enumerate(cards):
        x = 15 + i * (box_w + gap)
        pdf.set_fill_color(248, 250, 252); pdf.set_draw_color(226, 232, 240)
        pdf.rect(x, y, box_w, 24, "DF")
        pdf.set_text_color(100, 116, 139); pdf.set_font("Arial", "B", 6.8); pdf.set_xy(x, y + 4)
        pdf.cell(box_w, 4, safe_pdf_text(lab), align="C")
        pdf.set_text_color(30, 41, 59); pdf.set_font("Arial", "B", 15); pdf.set_xy(x, y + 10)
        pdf.cell(box_w, 8, safe_pdf_text(str(val)), align="C")
    pdf.set_y(y + 32)
    
    _pdf_section(pdf, "Executive Summary")
    pdf.set_x(15)
    pdf.set_text_color(52, 64, 84); pdf.set_font("Arial", "", 9.5)
    summary = (f"The selected report contains {k['projects']} project update(s) across {k['resources']} resource(s). "
               f"{k['updates']} project(s) include a weekly delivery update. Average completion is {k['completion']}% and average utilization is {k['utilization']}%. "
               f"There are {k['risks']} item(s) requiring risk or blocker attention.")
    pdf.multi_cell(180, 5.8, safe_pdf_text(summary))

    _pdf_section(pdf, "Project Status")
    _pdf_table(pdf, report["data"], ["Project", "Resource", "Status", "Comp.", "Util.", "Expected"], [60, 28, 28, 17, 17, 30], lambda r: [r.get("Project / Dashboard",""), r.get("Resource Name","") or r.get("Resource",""), r.get("Status",""), r.get("Completion %",""), r.get("Utilization %",""), r.get("Updated Expected Delivery date","")])
    
    _pdf_section(pdf, "Weekly Highlights")
    for p,t in report["highlights"]:
        pdf.set_x(15)
        pdf.set_text_color(22, 50, 79); pdf.set_font("Arial", "B", 9); pdf.multi_cell(180, 5, _pdf_safe(p))
        pdf.set_x(15)
        pdf.set_text_color(71, 84, 103); pdf.set_font("Arial", "", 9); pdf.multi_cell(180, 5, _pdf_safe(f"Delivered: {t}")); pdf.ln(1.5)
    if not report["highlights"]: 
        pdf.set_x(15)
        pdf.set_text_color(100, 116, 139); pdf.set_font("Arial", "", 9); pdf.multi_cell(180, 5, "No weekly accomplishment updates reported.")

    _pdf_section(pdf, "Risks & Blockers")
    if report["risks_df"].empty:
        pdf.set_x(15)
        pdf.set_text_color(22, 128, 91); pdf.set_font("Arial", "B", 9); pdf.multi_cell(180, 6, "No active blockers or risks reported for this period.")
    else:
        for _, row in report["risks_df"].iterrows():
            pdf.set_x(15)
            pdf.set_text_color(194, 65, 61); pdf.set_font("Arial", "B", 9); pdf.multi_cell(180, 5, _pdf_safe(row.get("Project / Dashboard", "N/A")))
            pdf.set_x(15)
            pdf.set_text_color(71, 84, 103); pdf.set_font("Arial", "", 9)
            blocker = row.get("Blocker", "") or f"Status marked as: {row.get('Status', 'N/A')}"
            pdf.multi_cell(180, 5, _pdf_safe(blocker)); pdf.ln(1.5)

    try: 
        out = pdf.output(dest="S"); return out.encode("latin-1") if isinstance(out, str) else bytes(out)
    except Exception: return bytes(pdf.output())

# =========================================================
# EXCEL REPORT GENERATOR
# =========================================================
def create_excel_report(dataframe, report_title):
    if not OPENPYXL_AVAILABLE:
        buffer=io.BytesIO();dataframe.to_excel(buffer,index=False);return buffer.getvalue()
    report=_build_report_model(dataframe,report_title);k=report["kpis"];wb=Workbook();buffer=io.BytesIO()
    
    ws = wb.active; ws.title = "Weekly Notes"
    cols = ["Resource", "Business Owner", "Project / Dashboard", "Start date", "Initial Delivery date", "Updated Expected Delivery date", "Work Type", "This Week Delivered", "Next Week Priority", "Completion %", "Utilization %", "Status", "Blocker"]
    
    ws["A1"] = report["title"]; ws["A1"].fill = PatternFill("solid", fgColor="FFD966"); ws["A1"].font = Font(bold=True, color="000000"); ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"] = f"{report['account_team_str']} ({report['generated'].strftime('%d %b %Y')})"; ws["B1"].fill = PatternFill("solid", fgColor="004B87"); ws["B1"].font = Font(bold=True, color="FFFFFF"); ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=len(cols)); ws.row_dimensions[1].height = 25
    
    header_fill = PatternFill("solid", fgColor="DDEBF7"); header_font = Font(bold=True, color="000000"); thin_border = Border(left=Side(style='thin', color='000000'), right=Side(style='thin', color='000000'), top=Side(style='thin', color='000000'), bottom=Side(style='thin', color='000000'))
    for c_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=2, column=c_idx, value=col_name); cell.fill = header_fill; cell.font = header_font; cell.border = thin_border; cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
    ws.row_dimensions[2].height = 35
    
    resource_fill = PatternFill("solid", fgColor="FCE4D6"); owner_fill = PatternFill("solid", fgColor="E2EFDA")
    for r_idx, row in enumerate(report["data"].to_dict('records'), 3):
        for c_idx, col_name in enumerate(cols, 1):
            val = row.get(col_name, ""); cell = ws.cell(row=r_idx, column=c_idx, value=val); cell.border = thin_border; cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_name == "Resource": cell.fill = resource_fill; cell.font = Font(bold=True, color="004B87")
            elif col_name == "Business Owner": cell.fill = owner_fill; cell.font = Font(bold=True, color="004B87")
            elif col_name == "Status": cell.font = Font(bold=True)
    widths = {"A": 16, "B": 14, "C": 30, "D": 11, "E": 11, "F": 13, "G": 22, "H": 38, "I": 35, "J": 12, "K": 12, "L": 14, "M": 38}
    for col_letter, width in widths.items(): ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"; ws.auto_filter.ref = f"A2:M{max(2,len(report['data'])+2)}"

    ex = wb.create_sheet("Executive Summary")
    metrics=[("Report Period",report["period"]),("Projects",k["projects"]),("Resources",k["resources"]),("Avg Completion",f"{k['completion']}%"),("Avg Utilization",f"{k['utilization']}%"),("Portfolio Health",f"{k['health']}/100"),("Healthy",sum(1 for _,r in report["data"].iterrows() if health_score(r)>=80)),("Attention",sum(1 for _,r in report["data"].iterrows() if 60<=health_score(r)<80)),("Critical",sum(1 for _,r in report["data"].iterrows() if health_score(r)<60))]
    ex.append(["Executive Summary",""]); ex.merge_cells("A1:B1"); ex["A1"].font=Font(size=16,bold=True,color="FFFFFF"); ex["A1"].fill=PatternFill("solid",fgColor="004B87")
    for kk,vv in metrics: ex.append([kk,vv])
    for cell in ex["A"]: cell.font=Font(bold=True)
    ex.column_dimensions["A"].width=25; ex.column_dimensions["B"].width=35; ex.freeze_panes="A2"

    ra = wb.create_sheet("Risks & Actions")
    ra.append(["Priority","Project / Dashboard","Account","Team","Resource","Status","Delivery","Reason","Recommended Action"])
    for item in report["actions"]: ra.append([item.get(k,"") for k in ["Priority","Project / Dashboard","Account","Team","Resource","Status","Delivery","Reason","Recommended Action"]])

    wow = wb.create_sheet("Week-over-Week")
    if report["movement"].empty: wow.append(["Select exactly one Year, Month and Week in the app to populate week-over-week movement."])
    else: wow.append(list(report["movement"].columns)); [wow.append(list(r)) for r in report["movement"].itertuples(index=False,name=None)]

    dq=wb.create_sheet("Data Quality")
    dq.append(["Project / Dashboard","Resource","Status","Issues"])
    for _,r in report["data"].iterrows():
        flags=data_quality_flags(r)
        if flags: dq.append([r.get("Project / Dashboard",""),r.get("Resource",""),r.get("Status",""),"; ".join(flags)])

    for sh in [ex,ra,wow,dq]:
        for cell in sh[1]: cell.font=Font(bold=True, color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="004B87"); cell.alignment=Alignment(wrap_text=True)
        for col in range(1, sh.max_column+1): sh.column_dimensions[get_column_letter(col)].width=min(45,max(14,sh.column_dimensions[get_column_letter(col)].width or 14))
        sh.freeze_panes="A2"
    wb.save(buffer); return buffer.getvalue()

def send_report_email(pdf_bytes, report_title, recipients, cc_recipients=None, excel_bytes=None, html_body=None):
    recipients, invalid_to = _validated_emails([x.strip() for x in recipients if x.strip()])
    cc_recipients, invalid_cc = _validated_emails([x.strip() for x in (cc_recipients or []) if x.strip()])
    if not recipients: return False, "Please enter at least one valid recipient email address."
    
    cfg = st.secrets if hasattr(st, "secrets") else {}
    host = cfg.get("SMTP_HOST", os.getenv("SMTP_HOST", ""))
    port = int(cfg.get("SMTP_PORT", os.getenv("SMTP_PORT", "587")))
    username = cfg.get("SMTP_USERNAME", os.getenv("SMTP_USERNAME", ""))
    password = cfg.get("SMTP_PASSWORD", os.getenv("SMTP_PASSWORD", ""))
    sender = cfg.get("SMTP_FROM", os.getenv("SMTP_FROM", username))
    use_tls = str(cfg.get("SMTP_USE_TLS", os.getenv("SMTP_USE_TLS", "true"))).lower() == "true"
    
    if not host or not sender: return False, "Email sending is not configured (SMTP secrets missing)."
    
    msg = EmailMessage()
    msg["Subject"] = report_title
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    if cc_recipients: msg["Cc"] = ", ".join(cc_recipients)
    
    if html_body:
        msg.set_content(f"Please find attached the {report_title}. To view the rich report, please enable HTML in your email client.")
        msg.add_alternative(html_body, subtype='html')
        
        if os.path.exists("logo.png") and "cid:factspan_logo" in html_body:
            try:
                with open("logo.png", "rb") as img:
                    msg.get_payload()[1].add_related(img.read(), maintype='image', subtype='png', cid='<factspan_logo>')
            except Exception: pass
    else:
        msg.set_content(f"""Hello,\n\nPlease find attached the {report_title}.\n\nThe report includes the executive portfolio view, delivery outlook, risks/actions, health signals and data-quality checks.\n\nGenerated by: {st.session_state.get('current_name','')} ({st.session_state.get('current_email','')})\nGenerated at: {datetime.now(IST).strftime('%d %b %Y, %I:%M %p')}\n\nRegards,\nWeekly Project Report""")
        
    if pdf_bytes:
        msg.add_attachment(pdf_bytes, maintype="application", subtype="pdf", filename="Weekly_Project_Report.pdf")
    if excel_bytes:
        msg.add_attachment(excel_bytes, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="Weekly_Project_Report.xlsx")
        
    try:
        with smtplib.SMTP(host, port, timeout=20) as server:
            if use_tls: server.starttls()
            if username and password: server.login(username, password)
            server.send_message(msg)
        return True, f"Report sent successfully to {', '.join(recipients)}."
    except Exception as exc: return False, f"Unable to send report: {exc}"

def _validated_emails(values):
    valid=[];invalid=[]
    for raw in values:
        addr=getaddresses([raw])[0][1].strip().lower()
        if addr and re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+",addr):valid.append(addr)
        elif raw.strip():invalid.append(raw.strip())
    return list(dict.fromkeys(valid)),list(dict.fromkeys(invalid))

# =========================================================
# DATA + GLOBAL TOP AREA
# =========================================================
df = pd.DataFrame(st.session_state.notes_db)
filtered_df = pd.DataFrame()

st.markdown(f"""
<div class="app-hero">
    <div style="position: absolute; right: 25px; top: 25px; font-size: 13px; font-weight: bold; background: rgba(255,255,255,0.15); padding: 5px 12px; border-radius: 20px;">👤 {html_escape(st.session_state.current_name)} · {html_escape(st.session_state.current_email)} · {html_escape(st.session_state.current_role)}</div>
    <div class="app-hero-eyebrow">Weekly Project Management</div>
    <div class="app-hero-title">Weekly Project Report</div>
    <div class="app-hero-subtitle">Track progress, priorities, utilization and delivery across your portfolio.</div>
</div>
""", unsafe_allow_html=True)

util_c1, util_c2, util_c3, util_c4 = st.columns([1.2, 1.4, 5.2, 1.8])
with util_c1:
    if st.button("🔄 Refresh Data", key="refresh_data_top", use_container_width=True):
        refresh_cloud_data()
        st.session_state.last_refresh_at = datetime.now(IST)
        st.rerun()
with util_c2:
    st.caption(f"Data loaded: {st.session_state.last_refresh_at.strftime('%d %b %Y, %I:%M %p')}")
with util_c4:
    if st.button("🔒 Logout", key="logout_top", use_container_width=True):
        st.session_state.authenticated = False
        st.session_state.current_email = ""
        st.session_state.current_name = ""
        st.session_state.current_role = ""
        st.session_state.current_scope = ""
        st.session_state.current_acc_scope = ""
        st.rerun()

# Navigation
labels = ["✍️ Enter Notes", "📋 Weekly Summary", "📈 Executive Report", "🚦 Action Center"]
cols = st.columns([1,1,1,1,1.5]) 
for col, label, m_idx in zip(cols[:4], labels, [1, 2, 3, 4]):
    with col:
        state = "active" if st.session_state.nav_selection == label else "inactive"
        st.markdown(f'<div class="marker-nav{m_idx}-{state}"></div>', unsafe_allow_html=True)
        if st.button(label, key=f"nav_{label}", use_container_width=True):
            st.session_state.nav_selection = label
            st.rerun()

nav_selection = st.session_state.nav_selection

if nav_selection == "📋 Weekly Summary":
    st.markdown('<div class="page-head"><div class="page-title">📋 Weekly Summary</div><div class="page-subtitle">Review project delivery, performance, utilization and current risks.</div></div>', unsafe_allow_html=True)
elif nav_selection == "📈 Executive Report":
    st.markdown('<div class="page-head"><div class="page-title">📈 Executive Report</div><div class="page-subtitle">Executive portfolio view of delivery, ownership, utilization and risks.</div></div>', unsafe_allow_html=True)
elif nav_selection == "🚦 Action Center":
    st.markdown('<div class="page-head"><div class="page-title">🚦 Action Center</div><div class="page-subtitle">Prioritized delivery risks, overdue work, data-quality issues and items needing management attention.</div></div>', unsafe_allow_html=True)

if nav_selection != "✍️ Enter Notes" and not df.empty:
    
    st.markdown("""
    <style>
    div[data-testid="stSelectbox"], div[data-testid="stMultiSelect"] {
        background: linear-gradient(to bottom, #ffffff, #f8fafc) !important; border: 1px solid #cbd5e1 !important;
        box-shadow: 0 3px 0 #cbd5e1, 0 4px 6px rgba(0,0,0,0.03) !important; border-radius: 12px !important;
        padding: 5px 8px 8px 8px !important; margin-bottom: 8px !important; transition: all 0.15s ease-out !important;
    }
    div[data-testid="stSelectbox"]:active, div[data-testid="stMultiSelect"]:active {
        transform: translateY(3px) !important; box-shadow: 0 0px 0 transparent, 0 0px 0px rgba(0,0,0,0) !important;
    }
    div[data-testid="stSelectbox"] [data-baseweb="select"] > div, div[data-testid="stMultiSelect"] [data-baseweb="select"] > div {
        background: #ffffff !important; border: 1px solid #cbd5e1 !important; box-shadow: inset 0 2px 4px rgba(0,0,0,0.02) !important;
        border-radius: 8px !important; min-height: 40px !important; transform: none !important;
    }
    </style>
    """, unsafe_allow_html=True)

    active_filter_count = 0
    if st.session_state.get("filter_resources"): active_filter_count += 1
    if st.session_state.get("filter_year"): active_filter_count += 1
    if st.session_state.get("filter_month"): active_filter_count += 1
    if st.session_state.get("filter_week"): active_filter_count += 1
    if st.session_state.get("quick_filter_select", "All") != "All": active_filter_count += 1
    filter_title = f"🔎 Filters & Search {'· ' + str(active_filter_count) + ' active' if active_filter_count else '· All records'}"
    
    with st.expander(filter_title, expanded=True):
        st.caption("Use the compact two-row filter bar. Selected multi-filters can be removed directly from their fields.")
        
        resource_options_all = list(df["Resource"].dropna().unique())
        
        def preset_my_updates_cb():
            st.session_state.filter_resources = [st.session_state.current_email] if st.session_state.current_email in resource_options_all else []
            st.session_state.quick_filter_select = "All"
        def preset_at_risk_cb(): st.session_state.quick_filter_select = "At Risk / Blocked"
        def preset_overdue_cb(): st.session_state.quick_filter_select = "Overdue"
        def preset_due_soon_cb(): st.session_state.quick_filter_select = "Due Soon"
        def preset_exec_view_cb():
            st.session_state.filter_resources=[]; st.session_state.filter_year=[]; st.session_state.filter_month=[]; st.session_state.filter_week=[]; st.session_state.quick_filter_select="All"; st.session_state.project_search_input=""
        def clear_filters_cb():
            for k,v in [("filter_account","All"),("filter_team","All"),("filter_resources",[]),("filter_year",[]),("filter_month",[]),("filter_week",[]),("quick_filter_select","All"),("project_search_input","")]:
                st.session_state[k]=v

        pv1,pv2,pv3,pv4,pv5 = st.columns(5)
        with pv1: st.button("My Updates", key="preset_my_updates", use_container_width=True, on_click=preset_my_updates_cb)
        with pv2: st.button("At Risk", key="preset_at_risk", use_container_width=True, on_click=preset_at_risk_cb)
        with pv3: st.button("Overdue", key="preset_overdue", use_container_width=True, on_click=preset_overdue_cb)
        with pv4: st.button("Due Soon", key="preset_due_soon", use_container_width=True, on_click=preset_due_soon_cb)
        with pv5: st.button("Executive View", key="preset_exec_view", use_container_width=True, on_click=preset_exec_view_cb)
        
        f_col1, f_col2, f_col3, f_col4 = st.columns(4)
        with f_col1:
            filter_acc = st.selectbox("Account", options=["All"] + list(df["Account"].dropna().unique()), key="filter_account")
            filtered_df = df if filter_acc == "All" else df[df["Account"] == filter_acc]
        with f_col2:
            filter_team = st.selectbox("Team", options=["All"] + list(filtered_df["Team"].dropna().unique()), key="filter_team")
            if filter_team != "All": filtered_df = filtered_df[filtered_df["Team"] == filter_team]
        with f_col3:
            res_options = list(filtered_df["Resource"].dropna().unique())
            selected_resources = st.multiselect("Resource", options=res_options, key="filter_resources", placeholder="All resources")
            if selected_resources: filtered_df = filtered_df[filtered_df["Resource"].isin(selected_resources)]
        with f_col4:
            project_search = st.text_input("Search Projects", key="project_search_input", placeholder="Project / dashboard name")
            if project_search: filtered_df = filtered_df[filtered_df["Project / Dashboard"].astype(str).str.contains(project_search, case=False, na=False)]

        f2_col1, f2_col2, f2_col3, f2_col4 = st.columns(4)
        with f2_col1:
            year_options = list(filtered_df["Year"].dropna().unique())
            filter_year = st.multiselect("Year", options=year_options, key="filter_year", placeholder="All years")
            if filter_year: filtered_df = filtered_df[filtered_df["Year"].isin(filter_year)]
        with f2_col2:
            month_options = list(filtered_df["Month"].dropna().unique())
            filter_month = st.multiselect("Month", options=month_options, key="filter_month", placeholder="All months")
            if filter_month: filtered_df = filtered_df[filtered_df["Month"].isin(filter_month)]
        with f2_col3:
            week_options = list(filtered_df["Week"].dropna().unique())
            filter_week = st.multiselect("Week", options=week_options, key="filter_week", placeholder="All weeks")
            if filter_week: filtered_df = filtered_df[filtered_df["Week"].isin(filter_week)]
        with f2_col4:
            quick_filters_opts = ["All", "At Risk / Blocked", "Due Soon", "Overdue", "Missing Update", "Data Quality Issues"]
            quick_filter = st.selectbox("Quick Filter", options=quick_filters_opts, key="quick_filter_select")
            if quick_filter == "At Risk / Blocked": filtered_df = filtered_df[filtered_df["Status"].isin(["At Risk", "Blocked"])]
            elif quick_filter == "Due Soon": filtered_df = filtered_df[filtered_df.apply(lambda r: delivery_state(r).startswith("Due in"), axis=1)]
            elif quick_filter == "Overdue": filtered_df = filtered_df[filtered_df.apply(lambda r: delivery_state(r).startswith("Overdue"), axis=1)]
            elif quick_filter == "Missing Update": filtered_df = filtered_df[filtered_df.apply(lambda r: "No weekly delivery update reported" in data_quality_flags(r), axis=1)]
            elif quick_filter == "Data Quality Issues": filtered_df = filtered_df[filtered_df.apply(lambda r: len(data_quality_flags(r)) > 0, axis=1)]

        rc1, rc2 = st.columns([1,7])
        with rc1:
            st.button("↺ Clear Filters", key="clear_all_filters", use_container_width=True, on_click=clear_filters_cb)
        with rc2:
            chips=[]
            if filter_acc != "All": chips.append(f"Account: {filter_acc}")
            if filter_team != "All": chips.append(f"Team: {filter_team}")
            if selected_resources: chips.append(f"Resources: {len(selected_resources)}")
            if filter_year: chips.append(f"Year: {format_badge_text(filter_year, year_options)}")
            if filter_month: chips.append(f"Month: {format_badge_text(filter_month, month_options)}")
            if filter_week: chips.append(f"Week: {format_badge_text(filter_week, week_options)}")
            if quick_filter != "All": chips.append(f"Quick: {quick_filter}")
            if project_search: chips.append(f"Search: {project_search}")
            st.markdown("<div class='filter-summary'>" + ("".join(f"<span class='filter-chip'>{html_escape(c)}</span>" for c in chips) if chips else "<span class='filter-caption'>No filters applied · showing all available records</span>") + "</div>", unsafe_allow_html=True)

    reporting_month = format_badge_text(filter_month, month_options)
    reporting_week = format_badge_text(filter_week, week_options)

    if not filtered_df.empty:
        if nav_selection == "📋 Weekly Summary":
            a1, a2, _ = st.columns([1.5, 1.5, 7])
            with a1:
                st.markdown('<div class="marker-send"></div>', unsafe_allow_html=True)
                if st.button("📤 Send Report", key="send_tab2", use_container_width=True):
                    st.session_state["show_send_dialog"] = "weekly"
            with a2:
                if FPDF_AVAILABLE:
                    pdf_bytes = create_pdf_report(filtered_df, f"Weekly Summary: {reporting_month} - {reporting_week}")
                    st.markdown('<div class="marker-pdf"></div>', unsafe_allow_html=True)
                    st.download_button("📄 Download PDF", data=pdf_bytes, file_name="Weekly_Summary.pdf", mime="application/pdf", key="pdf_tab2", use_container_width=True)
        elif nav_selection == "📈 Executive Report":
            a1, a2, a3, _ = st.columns([1.5, 1.5, 1.5, 5.5])
            with a1:
                st.markdown('<div class="marker-send"></div>', unsafe_allow_html=True)
                if st.button("📤 Send Report", key="send_tab3", use_container_width=True):
                    st.session_state["show_send_dialog"] = "executive"
            with a2:
                excel_bytes = create_excel_report(filtered_df, f"Executive Report: {reporting_month} - {reporting_week}")
                st.markdown('<div class="marker-excel"></div>', unsafe_allow_html=True)
                st.download_button("📊 Export Excel", data=excel_bytes, file_name="Weekly_Notes_Executive_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="excel_tab3", use_container_width=True)
            with a3:
                if FPDF_AVAILABLE:
                    pdf_bytes = create_pdf_report(filtered_df, f"Executive Report: {reporting_month} - {reporting_week}")
                    st.markdown('<div class="marker-pdf"></div>', unsafe_allow_html=True)
                    st.download_button("📄 Export PDF", data=pdf_bytes, file_name="Weekly_Notes_Executive_Report.pdf", mime="application/pdf", key="pdf_tab3", use_container_width=True)
else:
    if nav_selection == "✍️ Enter Notes": filtered_df = df

# =========================================================
# SEND REPORT DIALOG
# =========================================================
if st.session_state.get("show_send_dialog"):
    report_kind = st.session_state["show_send_dialog"]
    report_title = f"{'Weekly Summary' if report_kind == 'weekly' else 'Executive Report'}: {reporting_month} - {reporting_week}"
    with st.expander("📤 Send Report", expanded=True):
        st.markdown("**Send the generated PDF report by email**")
        to_value = st.text_input("To", placeholder="name@company.com, another@company.com", key="send_to")
        cc_value = st.text_input("CC (optional)", placeholder="name@company.com", key="send_cc")
        subject_value = st.text_input("Subject", value=report_title, key="send_subject")
        b1, b2 = st.columns([1, 1])
        with b1:
            if st.button("✉️ Send Now", type="primary", use_container_width=True, key="send_now"):
                to_list = [x.strip() for x in to_value.replace(";", ",").split(",") if x.strip()]
                cc_list = [x.strip() for x in cc_value.replace(";", ",").split(",") if x.strip()]
                with st.spinner("Generating leadership email, PDF and Excel report..."):
                    report_model = _build_report_model(filtered_df, subject_value)
                    current_pdf = create_pdf_report(filtered_df, subject_value) if FPDF_AVAILABLE else None
                    current_excel = create_excel_report(filtered_df, subject_value)
                    current_html = generate_html_report(report_model)
                if current_pdf is None: st.error("PDF generation unavailable.")
                else:
                    ok, message = send_report_email(current_pdf, subject_value, to_list, cc_list, current_excel, current_html)
                    if ok: st.success(message); st.session_state["show_send_dialog"] = None
                    else: st.error(message)
        with b2:
            if st.button("Cancel", use_container_width=True, key="cancel_send"):
                st.session_state["show_send_dialog"] = None

if st.session_state.show_success:
    st.toast(f"✅ {st.session_state.success_message}")
    st.session_state.show_success = False

# =========================================================
# VIEW 1: ENTER NOTES
# =========================================================
if nav_selection == "✍️ Enter Notes":
    st.markdown('<div class="page-head"><div class="page-title">✍️ Enter Weekly Notes</div><div class="page-subtitle">Capture your team\'s weekly progress, priorities, utilization and risks.</div></div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header">01 — Core Assignment</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col_acc, col_team = st.columns(2)
        with col_acc:
            account_options = ["➕ Add New Account..."] + list(st.session_state.accounts_db.keys())
            selected_account = st.selectbox("Account", options=account_options)
            final_account = st.text_input("Enter New Account Name:") if selected_account == "➕ Add New Account..." else selected_account
        with col_team:
            if final_account and final_account != "➕ Add New Account...":
                team_list = st.session_state.accounts_db.get(final_account, [])
                team_options = ["➕ Add New Team..."] + team_list
                selected_team = st.selectbox("Team", options=team_options)
                final_team = st.text_input("Enter New Team Name:") if selected_team == "➕ Add New Team..." else selected_team
            else:
                st.info("Select or enter an Account first to proceed.")
                final_team = None

    if final_account and final_team:
        st.markdown('<div class="section-header">02 — Timeframe & Ownership</div>', unsafe_allow_html=True)
        with st.container(border=True):
            t_col1, t_col2, t_col3, t_col4 = st.columns(4)
            with t_col1: year = st.selectbox("Year", options=[datetime.now(IST).year, datetime.now(IST).year+1, datetime.now(IST).year+2])
            with t_col2: month = st.selectbox("Month", options=["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"])
            with t_col3: week = st.selectbox("Week", options=["Week 1", "Week 2", "Week 3", "Week 4", "Week 5"])
            with t_col4: resource = st.text_input("Resource Email", value=st.session_state.current_email)
        st.caption(f"Reporting user: {html_escape(st.session_state.get('current_name',''))} · {html_escape(st.session_state.get('current_email',''))}")

        copy_col1, copy_col2, copy_col3 = st.columns([2.2, 2.2, 5.6])
        with copy_col1:
            if st.button("📋 Load Previous Week", key="load_previous_week", use_container_width=True):
                prev_source_df = pd.DataFrame(st.session_state.notes_db) if st.session_state.notes_db else pd.DataFrame()
                prev_for_copy = get_latest_prior_week(prev_source_df, year, month, week)
                if prev_for_copy.empty:
                    st.info("No previous reported week was found for this timeframe.")
                else:
                    same_resource = prev_for_copy[prev_for_copy["Resource"].astype(str).str.lower() == str(resource).strip().lower()]
                    if same_resource.empty:
                        st.info("No previous-week entries found for this resource.")
                    else:
                        st.session_state.copy_previous_candidates = same_resource.to_dict("records")
                        st.session_state.copy_source_loaded = True
                        st.session_state.project_count = max(1, len(same_resource))
                        st.rerun()
        with copy_col2:
            if st.session_state.get("copy_source_loaded"):
                if st.button("✖ Clear Copied Values", key="clear_copied_values", use_container_width=True):
                    st.session_state.copy_previous_records = []
                    st.session_state.copy_previous_candidates = []
                    st.session_state.copy_source_loaded = False
                    st.rerun()
        if st.session_state.get("copy_source_loaded"):
            candidates = st.session_state.get("copy_previous_candidates", [])
            st.markdown("**Choose projects to carry forward:**")
            selected_copy = []
            for ci, rec in enumerate(candidates):
                checked = st.checkbox(str(rec.get("Project / Dashboard", "Untitled")), value=True, key=f"copy_pick_{ci}")
                if checked: selected_copy.append(rec)
            st.session_state.copy_previous_records = selected_copy
            if st.button("Load Selected Projects", key="load_selected_previous", type="primary"):
                st.session_state.project_count = max(1, len(selected_copy)); st.rerun()
            st.caption("Previous-week values are loaded as a starting point only. Review dates, completion, delivery notes and blockers before saving.")

        st.markdown('<div class="section-header">03 — Weekly Accomplishments</div>', unsafe_allow_html=True)
        projects_to_save = []

        for i in range(st.session_state.project_count):
            with st.container(border=True):
                st.markdown(f'<div class="card-title">Project Entry #{i + 1}</div>', unsafe_allow_html=True)
                st.markdown('<hr>', unsafe_allow_html=True)
                r1c1, r1c2 = st.columns(2)
                copy_rec = st.session_state.get("copy_previous_records", [])[i] if i < len(st.session_state.get("copy_previous_records", [])) else {}
                project = r1c1.text_input("Project / Dashboard Name", value=str(copy_rec.get("Project / Dashboard", "")), placeholder="e.g., Q3 Sales Analysis", key=f"proj_{i}")
                business_owner = r1c2.text_input("Business Owner", value=str(copy_rec.get("Business Owner", "")), placeholder="e.g., Marketing Team", key=f"owner_{i}")
                
                r2c1, r2c2, r2c3 = st.columns(3)
                start_date = r2c1.date_input("Start Date", value=parse_date(copy_rec.get("Start date")) if copy_rec else None, key=f"start_{i}")
                initial_delivery = r2c2.date_input("Initial Delivery Date", value=parse_date(copy_rec.get("Initial Delivery date")) if copy_rec else None, key=f"init_{i}")
                updated_delivery = r2c3.date_input("Updated Expected Delivery", value=parse_date(copy_rec.get("Updated Expected Delivery date")) if copy_rec else None, key=f"upd_{i}")
                
                r3c1, r3c2, r3c3, r3c4 = st.columns(4)
                wt_options = ["Development", "Enhancement", "Adhoc", "Migration", "Other"]
                existing_copy_wt = str(copy_rec.get("Work Type", ""))
                wt_selection = r3c1.selectbox("Work Type", wt_options, index=wt_options.index(existing_copy_wt) if existing_copy_wt in wt_options else 0, key=f"wt_sel_{i}")
                if wt_selection == "Other": work_type = r3c1.text_input("Specify Work Type", placeholder="Enter work type...", key=f"work_{i}")
                else: work_type = wt_selection
                
                status_opts = ["On Track", "At Risk", "Blocked", "Completed", "Not Started", "On Hold", "Cancelled"]
                copy_status = str(copy_rec.get("Status", "On Track"))
                status = r3c2.selectbox("Status", status_opts, index=status_opts.index(copy_status) if copy_status in status_opts else 0, key=f"status_{i}")
                completion_pct = r3c3.number_input("Completion %", min_value=0, max_value=100, step=5, value=int(pct_value(copy_rec.get("Completion %", 0))), key=f"comp_{i}")
                utilization_pct = r3c4.number_input("Utilization %", min_value=0, max_value=100, step=5, value=int(pct_value(copy_rec.get("Utilization %", 0))), key=f"util_{i}")
                
                c_text1, c_text2 = st.columns(2)
                this_week = c_text1.text_area("This Week Delivered", value="" if copy_rec else "", placeholder="What was accomplished this week?", height=85, key=f"this_{i}")
                next_week = c_text2.text_area("Next Week Priority", value=str(copy_rec.get("Next Week Priority", "")), placeholder="What is the focus for next week?", height=85, key=f"next_{i}")
                blocker = st.text_area("Blocker / Risks", value="", placeholder="Detail any risks or blockers here. (Leave blank if none)", height=68, key=f"block_{i}")
                
                projects_to_save.append({
                    "id": str(uuid.uuid4()), "Project / Dashboard": project, "Business Owner": business_owner,
                    "Start date": str(start_date) if start_date else "N/A", "Initial Delivery date": str(initial_delivery) if initial_delivery else "N/A",
                    "Updated Expected Delivery date": str(updated_delivery) if updated_delivery else "N/A", "Work Type": work_type,
                    "Completion %": f"{completion_pct}%", "Utilization %": f"{utilization_pct}%", "Status": status,
                    "This Week Delivered": this_week, "Next Week Priority": next_week, "Blocker": blocker
                })

        col_add, col_rem, _ = st.columns([2, 2, 6])
        with col_add:
            if st.button("➕ Add Project", use_container_width=True):
                st.session_state.project_count += 1
                st.rerun()
        with col_rem:
            if st.session_state.project_count > 1:
                if st.button("🗑️ Remove Last", use_container_width=True):
                    st.session_state.project_count -= 1
                    st.rerun()

        st.divider()
        sub_col1, sub_col2, sub_col3 = st.columns([2, 3, 2])
        with sub_col2:
            st.markdown('<div class="marker-nav1-active"></div>', unsafe_allow_html=True)
            if st.button("💾 Save to Google Sheets", type="primary", use_container_width=True):
                if not resource: st.error("Please enter your Resource Email in Step 2 before saving.")
                else:
                    missing_names = [i + 1 for i, p in enumerate(projects_to_save) if not p["Project / Dashboard"]]
                    if missing_names: st.error(f"Please enter a Project Name for Project(s): {', '.join(map(str, missing_names))}")
                    else:
                        with st.spinner("Saving data to Google Sheets..."):
                            final_entries = []
                            for proj_data in projects_to_save:
                                full_entry = {"Account": final_account, "Team": final_team, "Year": year, "Month": month, "Week": week, "Resource": resource, "Resource Name": st.session_state.get("current_name", "")}
                                full_entry.update(proj_data)
                                final_entries.append(full_entry)
                            
                            existing_df = pd.DataFrame(st.session_state.notes_db) if st.session_state.notes_db else pd.DataFrame()
                            duplicate_keys = set(existing_df.apply(row_submission_key, axis=1)) if not existing_df.empty else set()
                            duplicate_entries = [e for e in final_entries if row_submission_key(e) in duplicate_keys]
                            if duplicate_entries and not st.session_state.get("allow_duplicate_save", False):
                                st.session_state.pending_entries = final_entries
                                st.session_state.duplicate_count = len(duplicate_entries)
                                st.warning(f"{len(duplicate_entries)} matching weekly submission(s) already exist for this resource/project/week. Choose an action below.")
                                d1,d2,d3 = st.columns(3)
                                with d1:
                                    if st.button("↻ Update Existing", key="dup_update", use_container_width=True):
                                        matches = existing_df[existing_df.apply(lambda r: row_submission_key(r) in {row_submission_key(e) for e in duplicate_entries}, axis=1)]
                                        for _, old in matches.iterrows():
                                            candidate = next((e for e in duplicate_entries if row_submission_key(e)==row_submission_key(old)), None)
                                            if candidate: update_existing_note_in_gsheets({**old.to_dict(), **candidate, "id": old.get("id")}); write_audit_event("UPDATE_FROM_DUPLICATE_CHECK", old.get("id"), old.to_dict(), candidate)
                                        st.session_state.notes_db = load_data_from_gsheets(); st.session_state.pending_entries=[]; st.rerun()
                                with d2:
                                    if st.button("＋ Save Anyway", key="dup_save", use_container_width=True):
                                        st.session_state.allow_duplicate_save = True; st.rerun()
                                with d3:
                                    if st.button("Cancel", key="dup_cancel", use_container_width=True):
                                        st.session_state.pending_entries=[]; st.session_state.allow_duplicate_save=False; st.rerun()
                                st.stop()
                            save_new_notes_to_gsheets(final_entries)
                            for entry in final_entries: write_audit_event("CREATE", entry.get("id"), None, entry)
                            st.session_state.allow_duplicate_save = False
                            st.session_state.notes_db = load_data_from_gsheets()
                            st.session_state.data_synced = True
                            st.session_state.last_refresh_at = datetime.now(IST)
                            st.session_state.success_message = f"Notes successfully saved to Google Sheets."
                            st.session_state.show_success = True
                            st.session_state.project_count = 1
                            st.rerun()

# =========================================================
# VIEW 2: WEEKLY SUMMARY
# =========================================================
elif nav_selection == "📋 Weekly Summary":
    if df.empty or filtered_df.empty:
        st.info("No data available to generate the summary. Please adjust filters or enter data.")
    else:
        total_projects = len(filtered_df)
        total_resources = filtered_df["Resource"].nunique()
        delivered_count = len(filtered_df[filtered_df["This Week Delivered"].fillna("").astype(str).str.strip() != ""])
        avg_comp = int((filtered_df["Completion %"].apply(lambda x: get_pct_decimal(x) * 100)).mean()) if not filtered_df.empty else 0
        avg_util = resource_utilization(filtered_df) if not filtered_df.empty else 0
        risk_items = filtered_df[(filtered_df["Status"].isin(["At Risk", "Blocked"])) | (filtered_df["Blocker"].fillna("").astype(str).str.strip() != "")]
        risk_count = len(risk_items)

        st.markdown('<div class="section-header">Executive Summary</div>', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="summary-box">
            This week's portfolio includes <strong>{total_projects}</strong> active work item(s), with <strong>{delivered_count}</strong> reporting a delivery update. 
            Average completion is <strong>{avg_comp}%</strong> and average utilization is <strong>{avg_util}%</strong>. 
            There are <strong>{risk_count}</strong> item(s) requiring risk or blocker attention.
        </div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="section-header">Portfolio Snapshot</div>', unsafe_allow_html=True)
        s1, s2, s3, s4, s5 = st.columns(5)
        with s1: st.markdown(render_metric_card("RESOURCES", total_resources), unsafe_allow_html=True)
        with s2: st.markdown(render_metric_card("PROJECTS", total_projects), unsafe_allow_html=True)
        with s3: st.markdown(render_metric_card("DELIVERED", delivered_count), unsafe_allow_html=True)
        with s4: st.markdown(render_metric_card("COMPLETION", f"{avg_comp}%"), unsafe_allow_html=True)
        with s5: st.markdown(render_metric_card("UTILIZATION", f"{avg_util}%"), unsafe_allow_html=True)

        hscore, hcounts = portfolio_health(filtered_df)
        hcolor = "#16a34a" if hscore >= 80 else ("#d97706" if hscore >= 60 else "#dc2626")
        overdue = sum(1 for _, r in filtered_df.iterrows() if delivery_bucket(r) == "Overdue")
        due7 = sum(1 for _, r in filtered_df.iterrows() if delivery_bucket(r) == "Next 7 Days")
        slippages = [delivery_slippage_days(r) for _, r in filtered_df.iterrows() if delivery_slippage_days(r) > 0]
        avg_slip = int(round(sum(slippages)/len(slippages))) if slippages else 0
        st.markdown(f"""<div class='health-strip'><div class='health-title'>Portfolio Health</div><div style='display:flex;align-items:center;gap:18px;margin-top:3px'><div class='health-score'><span class='health-dot' style='background:{hcolor}'></span>{hscore}/100</div><div style='font-size:13px;color:#475569'>Healthy <b>{hcounts.get('Healthy',0)}</b> · Attention <b>{hcounts.get('Attention',0)}</b> · Critical <b>{hcounts.get('Critical',0)}</b></div><div style='font-size:13px;color:#475569;margin-left:auto'>Overdue <b>{overdue}</b> · Due in 7 days <b>{due7}</b> · Avg slippage <b>{avg_slip}d</b></div></div></div>""", unsafe_allow_html=True)
        st.markdown('<div class="section-header">Delivery Outlook</div>', unsafe_allow_html=True)
        o1,o2,o3,o4,o5 = st.columns(5)
        buckets = Counter(delivery_bucket(r) for _,r in filtered_df.iterrows())
        for c,label in zip([o1,o2,o3,o4,o5],["OVERDUE","NEXT 7 DAYS","8–14 DAYS","15+ DAYS","DELIVERED"]):
            key_map={"OVERDUE":"Overdue","NEXT 7 DAYS":"Next 7 Days","8–14 DAYS":"8–14 Days","15+ DAYS":"15+ Days","DELIVERED":"Delivered"}
            with c: st.markdown(render_metric_card(label, buckets.get(key_map[label],0)), unsafe_allow_html=True)

        st.markdown('<div class="section-header">Project / Dashboard Status</div>', unsafe_allow_html=True)
        html_lines = [
            '<table class="custom-table">',
            '<thead><tr><th>Resource Email</th><th>Business Owner</th><th>Project / Dashboard</th><th>Expected Delivery</th><th>Completion</th><th>Utilization</th><th>Status</th></tr></thead>',
            '<tbody>'
        ]
        for _, row in filtered_df.iterrows():
            raw_status = row['Status']
            if raw_status == "On Track": pill = "status-on-track"
            elif raw_status == "At Risk": pill = "status-in-progress" 
            elif raw_status == "Blocked": pill = "status-blocked"
            elif raw_status == "Completed": pill = "status-completed"
            else: pill = "status-not-started"
            display_status = "In Progress" if raw_status == "At Risk" else raw_status
            row_html = f"<tr><td>{html_escape(row['Resource'])}</td><td>{html_escape(row['Business Owner'])}</td><td><span class='proj-name'>{html_escape(row['Project / Dashboard'])}</span><span class='work-type'>{html_escape(row['Work Type'])}</span></td><td>{html_escape(row['Updated Expected Delivery date'])}</td><td>{html_escape(row['Completion %'])}</td><td>{html_escape(row['Utilization %'])}</td><td><span class='status-pill {pill}'>{html_escape(display_status)}</span></td></tr>"
            html_lines.append(row_html)
        html_lines.append("</tbody></table>")
        st.markdown("".join(html_lines), unsafe_allow_html=True)

        st.markdown('<div class="section-header">What Changed?</div>', unsafe_allow_html=True)
        single_year = filter_year[0] if len(filter_year) == 1 else None
        single_month = filter_month[0] if len(filter_month) == 1 else None
        single_week = filter_week[0] if len(filter_week) == 1 else None
        prior_df = get_latest_prior_week(df, single_year, single_month, single_week) if single_year and single_month and single_week else pd.DataFrame()
        movement_df = compare_current_to_prior(filtered_df, prior_df)
        if movement_df.empty:
            st.info("Select exactly one Year, Month and Week to see project-level week-over-week movement.")
        else:
            mv1, mv2, mv3, mv4 = st.columns(4)
            with mv1: st.markdown(render_metric_card("NEW", int((movement_df["Movement"] == "New").sum())), unsafe_allow_html=True)
            with mv2: st.markdown(render_metric_card("IMPROVED", int((movement_df["Movement"] == "Improved").sum())), unsafe_allow_html=True)
            with mv3: st.markdown(render_metric_card("DECLINED", int((movement_df["Movement"] == "Declined").sum())), unsafe_allow_html=True)
            with mv4: st.markdown(render_metric_card("STABLE", int((movement_df["Movement"] == "Stable").sum())), unsafe_allow_html=True)
            improved_count = int((movement_df["Movement"] == "Improved").sum())
            declined = movement_df[movement_df["Movement"] == "Declined"]
            st.markdown(f"<div class='mini-insight'>📌 <b>Management insight:</b> {improved_count} item(s) improved, {len(declined)} declined and {int((movement_df['Movement']=='New').sum())} are new in the selected period.</div>", unsafe_allow_html=True)
            if not declined.empty:
                st.warning("Completion declines of 10+ points were detected in the current selection.")
                st.dataframe(declined[["Project / Dashboard","Resource","Completion Δ","Status Change"]], use_container_width=True, hide_index=True)

        st.markdown('<div class="section-header">Trend & Momentum</div>', unsafe_allow_html=True)
        tr1,tr2 = st.columns([1.5,6.5])
        with tr1: trend_periods=st.selectbox("History",[4,8,12],index=1,key="trend_periods",format_func=lambda x:f"Last {x} reported weeks")
        with tr2: st.caption("Historical performance uses the same weekly records already stored in Google Sheets; no new data source is introduced.")
        trend_df=historical_trend(df,trend_periods)
        if trend_df.empty: st.info("Not enough historical weekly data for a trend view.")
        else: st.dataframe(trend_df,use_container_width=True,hide_index=True)

        st.markdown('<div class="section-header">Commitment vs Actual</div>', unsafe_allow_html=True)
        positive_slip=[delivery_slippage_days(r) for _,r in filtered_df.iterrows() if delivery_slippage_days(r)>0]
        slip1,slip2,slip3=st.columns(3)
        with slip1: st.markdown(render_metric_card("REVISED DELIVERY", len(positive_slip)), unsafe_allow_html=True)
        with slip2: st.markdown(render_metric_card("AVG SLIPPAGE", f"{int(round(sum(positive_slip)/len(positive_slip))) if positive_slip else 0}d"), unsafe_allow_html=True)
        with slip3: st.markdown(render_metric_card("MAX SLIPPAGE", f"{max(positive_slip) if positive_slip else 0}d"), unsafe_allow_html=True)

        st.markdown('<div class="section-header">Blockers & Risks</div>', unsafe_allow_html=True)
        if risk_count == 0: st.success("No active blockers or risks reported this week! 🎉")
        else:
            for _, row in risk_items.iterrows():
                blocker_text = row['Blocker'] if row['Blocker'] else f"Status marked as: {row['Status']}"
                st.markdown(f'<div class="risk-box"><div class="risk-title">⚠️ {row["Project / Dashboard"]}</div><div class="risk-desc">{blocker_text}</div></div>', unsafe_allow_html=True)

# =========================================================
# VIEW 3: EXECUTIVE REPORT (WITH SMART EDIT BUTTONS)
# =========================================================
elif nav_selection == "📈 Executive Report":
    if df.empty or filtered_df.empty:
        st.info("No data available. Go to the 'Enter Notes' tab to log your first update.")
    else:
        st.markdown('<div class="section-header">Portfolio Management View</div>', unsafe_allow_html=True)
        hscore, hcounts = portfolio_health(filtered_df)
        ec1, ec2, ec3, ec4 = st.columns([2, 2, 2, 4])
        with ec1: st.caption(f"{len(filtered_df)} project update(s) in current selection")
        with ec2: st.caption(f"{sum(1 for _, r in filtered_df.iterrows() if classify_attention(r)[0] == 'Critical')} critical item(s)")
        with ec3: st.caption(f"Portfolio health: {hscore}/100")
        with ec4:
            compact_mode = st.toggle("Compact executive view", value=st.session_state.get("executive_compact", False), key="executive_compact_toggle", help="Switch between the detailed project cards and a dense management table.")
            st.session_state.executive_compact = compact_mode
        if st.session_state.executive_compact:
            compact_rows=[]
            for _, r in filtered_df.iterrows():
                hs=health_score(r)
                compact_rows.append({"Account":r.get("Account",""),"Team":r.get("Team",""),"Project":r.get("Project / Dashboard",""),"Owner":r.get("Business Owner",""),"Status":r.get("Status",""),"Completion":r.get("Completion %",""),"Delivery":delivery_state(r),"Health":f"{hs}/100 · {health_label(hs)}","Risk":suggested_action(r) if classify_attention(r)[0] != "Normal" else "—"})
            st.dataframe(pd.DataFrame(compact_rows), use_container_width=True, hide_index=True)
        else:
          for account_name in filtered_df["Account"].dropna().unique():
            st.markdown(f"<div class='acc-header'>🏢 {account_name}</div>", unsafe_allow_html=True)
            account_data = filtered_df[filtered_df["Account"] == account_name]
            for team_name in account_data["Team"].dropna().unique():
                st.markdown(f"<div class='team-header'>{team_name}</div>", unsafe_allow_html=True)
                team_data = account_data[account_data["Team"] == team_name]
                for resource_name in team_data["Resource"].dropna().unique():
                    resource_data = team_data[team_data["Resource"] == resource_name]
                    with st.expander(f"👤 {resource_name} ({len(resource_data)} updates)", expanded=True):
                        for index, row in resource_data.iterrows():
                            rec_id = row.get('id')
                            if st.session_state.edit_id == rec_id:
                                with st.container(border=True):
                                    st.markdown(f"<div class='edit-banner'>✏️ Editing: {html_escape(row.get('Project / Dashboard',''))}</div>", unsafe_allow_html=True)
                                    e_r1c1, e_r1c2, e_r1c3 = st.columns(3)
                                    e_proj = e_r1c1.text_input("Project Name", value=row.get('Project / Dashboard', ''), key=f"ep_{rec_id}")
                                    e_bo = e_r1c2.text_input("Business Owner", value=row.get('Business Owner', ''), key=f"ebo_{rec_id}")
                                    wt_options = ["Development", "Enhancement", "Adhoc", "Migration", "Other"]
                                    existing_wt = row.get('Work Type', '')
                                    wt_idx = wt_options.index(existing_wt) if existing_wt in wt_options else 4
                                    e_wt_sel = e_r1c3.selectbox("Work Type", wt_options, index=wt_idx, key=f"ewtsel_{rec_id}")
                                    if e_wt_sel == "Other": e_wt = e_r1c3.text_input("Specify Work Type", value=existing_wt if existing_wt not in wt_options else "", key=f"ewt_{rec_id}")
                                    else: e_wt = e_wt_sel
                                        
                                    e_r2c1, e_r2c2, e_r2c3 = st.columns(3)
                                    e_start = e_r2c1.date_input("Start Date", value=parse_date(row.get('Start date')), key=f"esd_{rec_id}")
                                    e_init = e_r2c2.date_input("Initial Delivery", value=parse_date(row.get('Initial Delivery date')), key=f"eid_{rec_id}")
                                    e_upd = e_r2c3.date_input("Updated Delivery", value=parse_date(row.get('Updated Expected Delivery date')), key=f"eud_{rec_id}")
                                    e_r3c1, e_r3c2, e_r3c3 = st.columns(3)
                                    status_opts = ["On Track", "At Risk", "Blocked", "Completed", "Not Started", "On Hold", "Cancelled"]
                                    curr_status = row.get('Status', 'On Track')
                                    if curr_status not in status_opts: curr_status = "On Track"
                                    e_status = e_r3c1.selectbox("Status", status_opts, index=status_opts.index(curr_status), key=f"estat_{rec_id}")
                                    e_comp = e_r3c2.number_input("Completion %", min_value=0, max_value=100, step=5, value=int(pct_value(row.get('Completion %', '0'))), key=f"ecomp_{rec_id}")
                                    e_util = e_r3c3.number_input("Utilization %", min_value=0, max_value=100, step=5, value=int(pct_value(row.get('Utilization %', '0'))), key=f"eutil_{rec_id}")
                                    e_tw = st.text_area("This Week Delivered", value=str(row.get('This Week Delivered', '')), height=85, key=f"etw_{rec_id}")
                                    e_nw = st.text_area("Next Week Priority", value=str(row.get('Next Week Priority', '')), height=85, key=f"enw_{rec_id}")
                                    e_block = st.text_area("Blocker / Risks", value=str(row.get('Blocker', '')), height=68, key=f"eb_{rec_id}")
                                    st.divider()
                                    e_btn_c1, e_btn_c2, _ = st.columns([2, 2, 6])
                                    with e_btn_c1:
                                        if st.button("💾 Save Changes", type="primary", key=f"esave_{rec_id}", use_container_width=True):
                                            if not user_can_edit_row(row):
                                                st.error("Authorization failed. This record was not changed.")
                                            else:
                                                updated_data = {
                                                    "id": rec_id, "Account": account_name, "Team": team_name, "Year": row.get("Year"), "Month": row.get("Month"), "Week": row.get("Week"), "Resource": resource_name, "Resource Name": row.get("Resource Name", st.session_state.get("current_name", "")),
                                                    "Project / Dashboard": e_proj, "Business Owner": e_bo, "Work Type": e_wt,
                                                    "Start date": str(e_start) if e_start else "N/A", "Initial Delivery date": str(e_init) if e_init else "N/A", "Updated Expected Delivery date": str(e_upd) if e_upd else "N/A",
                                                    "Status": e_status, "Completion %": f"{e_comp}%", "Utilization %": f"{e_util}%",
                                                    "This Week Delivered": e_tw, "Next Week Priority": e_nw, "Blocker": e_block
                                                }
                                                with st.spinner("Updating Google Sheets..."):
                                                    update_existing_note_in_gsheets(updated_data)
                                                    write_audit_event("UPDATE", rec_id, row.to_dict(), updated_data)
                                                    st.session_state.notes_db = load_data_from_gsheets()
                                                    st.session_state.data_synced = True
                                                    st.session_state.last_refresh_at = datetime.now(IST)
                                                st.session_state.edit_id = None
                                                st.session_state.show_success = True
                                                st.session_state.success_message = "Cloud update saved successfully."
                                                st.rerun()
                                    with e_btn_c2:
                                        if st.button("Cancel", key=f"ecancel_{rec_id}", use_container_width=True):
                                            st.session_state.edit_id = None; st.rerun()
                            else:
                                with st.container(border=True):
                                    h_col1, h_col2 = st.columns([9, 1])
                                    with h_col1:
                                        st.markdown(f"<div class='card-title'> {html_escape(row.get('Project / Dashboard',''))}</div>", unsafe_allow_html=True)
                                        st.markdown(f"<span style='font-size:14px; color:#667085;'><b>Owner:</b> {html_escape(row.get('Business Owner',''))} &nbsp;|&nbsp; <b>Type:</b> {html_escape(row.get('Work Type',''))}</span>", unsafe_allow_html=True)
                                    with h_col2:
                                        if user_can_edit_row(row):
                                            if st.button("✏️ Edit", key=f"edit_btn_{rec_id}", use_container_width=True):
                                                st.session_state.edit_id = rec_id; st.rerun()
                                        else:
                                            st.markdown("<div style='color:#94a3b8; font-size:12px; text-align:center; padding-top:10px;'>🔒 View Only</div>", unsafe_allow_html=True)

                                    c1, c2, c3 = st.columns([2, 2, 1])
                                    with c1:
                                        st.markdown("<div style='font-size:13px; font-weight:800; color:#344054;'>THIS WEEK DELIVERED</div>", unsafe_allow_html=True)
                                        st.markdown(f"<div style='font-size:15px; color:#344054;'>{html_escape(row.get('This Week Delivered','')) if str(row.get('This Week Delivered','')).strip() else '<i>None reported.</i>'}</div>", unsafe_allow_html=True)
                                        st.markdown(f"<div style='font-size:13px; margin-top:7px; color:#667085;'>Start: <b>{html_escape(row.get('Start date',''))}</b></div>", unsafe_allow_html=True)
                                    with c2:
                                        st.markdown("<div style='font-size:13px; font-weight:800; color:#344054;'>NEXT WEEK PRIORITY</div>", unsafe_allow_html=True)
                                        st.markdown(f"<div style='font-size:15px; color:#344054;'>{html_escape(row.get('Next Week Priority','')) if str(row.get('Next Week Priority','')).strip() else '<i>None reported.</i>'}</div>", unsafe_allow_html=True)
                                        st.markdown(f"<div style='font-size:13px; margin-top:7px; color:#667085;'>Expected Delivery: <b>{html_escape(row.get('Updated Expected Delivery date',''))}</b></div>", unsafe_allow_html=True)
                                    with c3:
                                        st.markdown("<div style='font-size:13px; font-weight:800; color:#344054;'>METRICS</div>", unsafe_allow_html=True)
                                        st.markdown(f"<div style='font-size:14px; display:flex; justify-content:space-between;'><span>Completion</span><b>{html_escape(row.get('Completion %',''))}</b></div>", unsafe_allow_html=True)
                                        st.progress(get_pct_decimal(row.get('Completion %','')))
                                        st.markdown(f"<div style='font-size:14px; display:flex; justify-content:space-between; margin-top:3px;'><span>Utilization</span><b>{html_escape(row.get('Utilization %',''))}</b></div>", unsafe_allow_html=True)
                                        st.progress(get_pct_decimal(row.get('Utilization %','')))
                                    
                                    d_state = delivery_state(row)
                                    d_flags = data_quality_flags(row)
                                    if d_state.startswith("Overdue"):
                                        st.markdown(f"<div class='risk-box' style='margin-top:10px'><div class='risk-title'>⏰ {html_escape(d_state)}</div><div class='risk-desc'>Expected delivery: {html_escape(row.get('Updated Expected Delivery date','N/A'))}</div></div>", unsafe_allow_html=True)
                                    elif d_state.startswith("Due in"):
                                        st.markdown(f"<div class='edit-banner' style='margin-top:10px'>📅 <b>{html_escape(d_state)}</b> · Expected delivery: {html_escape(row.get('Updated Expected Delivery date','N/A'))}</div>", unsafe_allow_html=True)
                                    if d_flags:
                                        st.markdown(f"<div class='edit-banner' style='margin-top:10px'>🧹 <b>Data quality:</b> {html_escape('; '.join(d_flags))}</div>", unsafe_allow_html=True)

                                    if str(row.get("Blocker","")).strip().lower() not in ["none", "na", "n/a", ""]:
                                        st.markdown(f"<div class='alert-card alert-danger' style='margin-top:14px; margin-bottom:0; background:#fef2f2; border-left:4px solid #dc2626; padding:10px;'><div class='alert-title' style='font-weight:bold; color:#b91c1c;'>⚠ Blocker</div><div class='alert-desc' style='font-size:13px; color:#450a0a;'>{html_escape(row.get('Blocker',''))}</div></div>", unsafe_allow_html=True)

# =========================================================
# VIEW 4: ACTION CENTER
# =========================================================
elif nav_selection == "🚦 Action Center":
    if df.empty:
        st.info("No data available. Go to the 'Enter Notes' tab to log your first update.")
    else:
        action_items = build_action_items(filtered_df if not filtered_df.empty else df)
        critical = [x for x in action_items if x["Priority"] == "Critical"]
        attention = [x for x in action_items if x["Priority"] == "Attention"]
        a1, a2, a3, a4 = st.columns(4)
        with a1: st.markdown(render_metric_card("CRITICAL", len(critical)), unsafe_allow_html=True)
        with a2: st.markdown(render_metric_card("ATTENTION", len(attention)), unsafe_allow_html=True)
        with a3: st.markdown(render_metric_card("OVERDUE", sum(1 for x in action_items if str(x["Delivery"]).startswith("Overdue"))), unsafe_allow_html=True)
        with a4: st.markdown(render_metric_card("TOTAL FLAGGED", len(action_items)), unsafe_allow_html=True)
        
        st.markdown('<div class="section-header">Prioritized Work Items</div>', unsafe_allow_html=True)
        if not action_items:
            st.success("Everything is currently within the defined attention rules. 🎉")
        else:
            action_df = pd.DataFrame(action_items)
            st.dataframe(action_df[["Priority","Project / Dashboard","Account","Team","Resource","Status","Delivery","Reason","Recommended Action"]], use_container_width=True, hide_index=True)
            
        st.markdown('<div class="section-header">Resource Health & Workload</div>', unsafe_allow_html=True)
        resource_rows=[]
        for resource, grp in (filtered_df if not filtered_df.empty else df).groupby("Resource"):
            avg_c=int(round(sum(pct_value(x) for x in grp["Completion %"])/max(1,len(grp))))
            avg_u=int(round(sum(pct_value(x) for x in grp["Utilization %"])/max(1,len(grp))))
            criticals=sum(1 for _,r in grp.iterrows() if classify_attention(r)[0]=="Critical")
            workload="High" if avg_u >= 85 else ("Low" if avg_u < 50 else "Balanced")
            resource_rows.append({"Resource":resource,"Projects":len(grp),"Avg Completion":f"{avg_c}%","Avg Utilization":f"{avg_u}%","Critical":criticals,"Workload":workload})
        if resource_rows: st.dataframe(pd.DataFrame(resource_rows), use_container_width=True, hide_index=True)

        st.markdown('<div class="section-header">Submission Compliance</div>', unsafe_allow_html=True)
        scope_df = filtered_df if not filtered_df.empty else df
        if not scope_df.empty:
            compliance = scope_df.groupby("Resource").size().reset_index(name="Updates")
            compliance["Coverage"] = compliance["Updates"].apply(lambda x: "Reported" if x else "Missing")
            st.caption(f"Reporting coverage: {len(compliance)} resource(s) with at least one submitted update in the current selection.")
            st.dataframe(compliance, use_container_width=True, hide_index=True)

        st.markdown('<div class="section-header">Data Quality Checks</div>', unsafe_allow_html=True)
        dq_records = []
        for _, row in (filtered_df if not filtered_df.empty else df).iterrows():
            flags = data_quality_flags(row)
            if flags:
                dq_records.append({"Project / Dashboard": row.get("Project / Dashboard",""), "Resource": row.get("Resource",""), "Status": row.get("Status",""), "Issues": "; ".join(flags)})
        if dq_records:
            st.dataframe(pd.DataFrame(dq_records), use_container_width=True, hide_index=True)
        else:
            st.success("No data-quality exceptions detected.")
